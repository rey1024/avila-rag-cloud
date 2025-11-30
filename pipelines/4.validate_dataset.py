# ==========================================================
# DATASET VALIDATION PIPELINE (AVILA Version)
# GPT-4o Validator
# ==========================================================

import os
import json
import re
import pandas as pd
from tqdm import tqdm
from datetime import datetime
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import sys
from dotenv import load_dotenv
# memastikan folder project menjadi root import
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)
from utils_common.utils import *
import logging

# ----------------------------------------------------------
# 1. Konfigurasi
# ----------------------------------------------------------
load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

INPUT_EXCEL = "datasets/2.qa/ds_answer.xlsx"   # <=== ganti sesuai data anda
OUTPUT_DIR = f"results/output_validation/validation_{timestamp}"
os.makedirs(OUTPUT_DIR, exist_ok=True)
setup_logging(OUTPUT_DIR)
logger = logging.getLogger(__name__)
output_jsonl_path = f"{OUTPUT_DIR}/validation_{timestamp}.jsonl"


# ----------------------------------------------------------
# 2. Ekstraktor JSON robust
# ----------------------------------------------------------
def safe_json_extract(text):
    text = text.replace("```json", "").replace("```", "").strip()

    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        json_str = match.group(0)
        json_str = json_str.replace("\n", " ").replace("\r", " ")
        json_str = re.sub(r"[\x00-\x1F\x7F]", "", json_str)
        try:
            return json.loads(json_str)
        except:
            return None

    return None

# ---------------------------------------------
# PEMETAAN BLOOM LEVEL ANGKA → LABEL
# ---------------------------------------------
bloom_label_map = {
    1: "1. Remember",
    2: "2. Understand",
    3: "3. Apply",
    4: "4. Analyze",
    5: "5. Evaluate",
    6: "6. Create"
}

def get_bloom_label(value):
    try:
        v = int(value)
        return bloom_label_map.get(v, f"{value}")
    except:
        return str(value)

# ----------------------------------------------------------
# 3. Prompt Validator
# ----------------------------------------------------------

VALIDATOR_PROMPT = r"""
Anda bertugas sebagai validator kualitas dataset tanya jawab berbasis konteks.

Lakukan empat pemeriksaan berikut secara ketat:

1. Kesesuaian Pertanyaan dengan Konteks
   - Apakah pertanyaan bisa dijawab oleh konteks.
   - Jika tidak, beri label "Tidak Sesuai" dan jelaskan.

2. Kesesuaian Level Taksonomi Bloom
   - Cocokkan pertanyaan dengan level Bloom (Remember, Understand, Apply, Analyze, Evaluate, Create).
   - Jika tidak tepat, beri label "Tidak Sesuai" dan alasan.

3. Kesesuaian Jawaban dengan Pertanyaan
   - Apakah jawaban benar-benar menjawab pertanyaan.
   - Jika tidak, beri label "Tidak Sesuai" dan alasan.

4. Kesesuaian Jawaban dengan Konteks
   - Apakah jawaban sesuai fakta di konteks.
   - Jika tidak, beri label "Tidak Sesuai" dan alasan.

Format output WAJIB:
{
  "pertanyaan_vs_konteks": "Sesuai / Tidak Sesuai + alasan jika tidak",
  "level_bloom_vs_pertanyaan": "Sesuai / Tidak Sesuai + alasan jika tidak",
  "jawaban_vs_pertanyaan": "Sesuai / Tidak Sesuai + alasan jika tidak",
  "jawaban_vs_konteks": "Sesuai / Tidak Sesuai + alasan jika tidak"
}
"""

# ----------------------------------------------------------
# 4. Fungsi panggilan GPT Validator
# ----------------------------------------------------------

def validate_row(context, question, bloom_level, answer):

    user_prompt = f"""
Konteks:
{context}

Pertanyaan:
{question}

Level Bloom:
{bloom_level}

Jawaban:
{answer}

Berikan hasil validasi dalam format JSON yang sudah ditentukan.
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": VALIDATOR_PROMPT},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.0,
            max_tokens=400
        )

        raw = response.choices[0].message.content.strip()
        data = safe_json_extract(raw)

        if data is None:
            return {
                "pertanyaan_vs_konteks": "Tidak Sesuai - Model tidak mengembalikan JSON valid",
                "level_bloom_vs_pertanyaan": "Tidak Sesuai - JSON rusak",
                "jawaban_vs_pertanyaan": "Tidak Sesuai - JSON rusak",
                "jawaban_vs_konteks": "Tidak Sesuai - JSON rusak",
            }

        return data

    except Exception as e:
        logger.info(f"❌ Error: {e}")
        return {
            "pertanyaan_vs_konteks": "Tidak Sesuai - Exception",
            "level_bloom_vs_pertanyaan": "Tidak Sesuai - Exception",
            "jawaban_vs_pertanyaan": "Tidak Sesuai - Exception",
            "jawaban_vs_konteks": "Tidak Sesuai - Exception",
        }


# ----------------------------------------------------------
# 5. Muat Dataset Excel
# ----------------------------------------------------------
df = pd.read_excel(INPUT_EXCEL)
#df = df.head(5) 

logger.info(f"📘 Loaded {len(df)} records from Excel")


# ----------------------------------------------------------
# 6. Proses Validasi
# ----------------------------------------------------------
results = []
output_jsonl = open(output_jsonl_path, "w", encoding="utf-8")

logger.info("🚀 Validating dataset...")

for idx, row in tqdm(df.iterrows(), total=len(df), desc="Validating"):

    context = str(row["context_groundtruth"])
    question = str(row["question"])
    #bloom = str(row["bloom_level"])
    bloom_numeric = row["bloom_level"]
    bloom = get_bloom_label(bloom_numeric)

    answer = str(row["answer_groundtruth"])

    validation = validate_row(context, question, bloom, answer)

    rec = {
        "context": context,
        "question": question,
        "bloom_level": bloom,
        "answer": answer,
        "validation": validation
    }

    results.append(rec)
    output_jsonl.write(json.dumps(rec, ensure_ascii=False) + "\n")

output_jsonl.close()


# ----------------------------------------------------------
# 7. Simpan Excel Output
# ----------------------------------------------------------
df_out = pd.DataFrame([
    {
        "context_text": r["context"],
        "question": r["question"],
        "bloom_level": r["bloom_level"],
        "answer_groundtruth": r["answer"],
        "Pertanyaan_vs_Konteks": r["validation"]["pertanyaan_vs_konteks"],
        "LevelBloom_vs_Pertanyaan": r["validation"]["level_bloom_vs_pertanyaan"],
        "Jawaban_vs_Pertanyaan": r["validation"]["jawaban_vs_pertanyaan"],
        "Jawaban_vs_Konteks": r["validation"]["jawaban_vs_konteks"],
    }
    for r in results
])

xlsx_path = f"{OUTPUT_DIR}/validation_{timestamp}.xlsx"
df_out.to_excel(xlsx_path, index=False)

# auto formatting
try:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    font = Font(name="Calibri", size=11)
    align = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = align

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    wb.save(xlsx_path)
    logger.info(f"✅ Excel formatted: {xlsx_path}")

except Exception as e:
    logger.info(f"⚠️ Formatting failed: {e}")


# ----------------------------------------------------------
# 8. Laporan Akhir
# ----------------------------------------------------------
logger.info("\n🎉 SELESAI VALIDASI DATASET")
logger.info(f"📁 Output Folder: {OUTPUT_DIR}")
logger.info(f"📄 JSONL: {output_jsonl_path}")
logger.info(f"📊 Total record divalidasi: {len(results)}")
