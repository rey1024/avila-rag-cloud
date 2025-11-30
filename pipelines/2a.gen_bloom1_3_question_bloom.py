# ==========================================================
# Generate Multi-Concept Bloom QUESTIONS Only (Fixed Version)
# ==========================================================

import os
import re
import json
import sys
import os

# memastikan folder project menjadi root import
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)

import pandas as pd
from dotenv import load_dotenv
from tqdm import tqdm
from datetime import datetime
from openai import OpenAI
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from utils_common.utils import setup_logging, log_var

# ----------------------------------------------------------
# 1. Konfigurasi
# ----------------------------------------------------------
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # ✅ DIPERBAIKI
load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
output_dir = f"results/ds_question/ds1-3_{timestamp}"
os.makedirs(output_dir, exist_ok=True)

import logging
setup_logging(output_dir)
logger = logging.getLogger(__name__)


input_file = "datasets/chunks_20251123_120207.xlsx"






logger.info(f"📘 Memuat dataset dari: {input_file}")
df = pd.read_excel(input_file)
#df = df.head(1)
total_rows = len(df)

logger.info(f"Total konsep: {total_rows}")

# ----------------------------------------------------------
# 2. Embedding Model
# ----------------------------------------------------------
# embedder = SentenceTransformer("models/embedding/multilingual-e5-base")
# df["embedding"] = df["chunk_text"].apply(lambda x: embedder.encode(str(x)))


# ----------------------------------------------------------
# NEW FUNCTION: Adaptive Multi-Context Selection for Bloom 4–6
# ----------------------------------------------------------
# def pick_related_concepts_adaptive(concept_id, concept_text, df, min_sim=0.25, high_sim=0.40):
#     """
#     Menghasilkan 2 atau 3 context sesuai similarity.
#     - Selalu ada primary context
#     - Tambahkan context kedua jika similarity > min_sim
#     - Tambahkan context ketiga hanya jika context kedua juga sangat mirip
#     """

#     this_emb = embedder.encode(concept_text)
#     sims = []

#     for _, row in df.iterrows():
#         cid = row["id"]
#         if cid == concept_id:
#             continue

#         sim = util.cos_sim(this_emb, row["embedding"]).item()
#         sims.append((cid, row["context"], sim))

#     # Urutkan berdasarkan similarity
#     sims_sorted = sorted(sims, key=lambda x: x[2], reverse=True)

#     # Filter kandidat sesuai ambang minimal
#     filtered = [(cid, ctx, sim) for cid, ctx, sim in sims_sorted if sim >= min_sim]

#     # Jika hanya primary context → gagal
#     if len(filtered) == 0:
#         return [concept_id]   # hanya primary (tidak ideal untuk Bloom 4–6)

#     # Selalu ambil primary + 1 related context
#     selected = [concept_id, filtered[0][0]]

#     # Tambah context ketiga jika similarity kandidat kedua cukup tinggi
#     if len(filtered) > 1 and filtered[1][2] >= high_sim:
#         selected.append(filtered[1][0])

#     return selected


# # ----------------------------------------------------------
# # 3. Pilih konsep terkait - DIPERBAIKI
# # ----------------------------------------------------------
# def pick_related_concepts(concept_id, concept_text, df, top_k=2):
#     this_emb = embedder.encode(concept_text)

#     sims = []
#     for _, row in df.iterrows():
#         cid = row["id"]
#         if cid == concept_id:
#             continue

#         sim = util.cos_sim(this_emb, row["embedding"]).item()
#         sims.append((cid, row["context"], sim))

#     sims_sorted = sorted(sims, key=lambda x: x[2], reverse=True)
    
#     # ✅ DIPERBAIKI: tuple unpacking yang konsisten
#     threshold = 0.25
#     filtered_sims = [(cid, ctx, sim) for cid, ctx, sim in sims_sorted if sim > threshold]
    
#     selected = [concept_id] + [cid for cid, _, _ in filtered_sims[:top_k]]
#     return selected

# ----------------------------------------------------------
# 4. Bloom Validator
# ----------------------------------------------------------
bloom_keywords = {
    "Remember": ["apa", "siapa", "kapan", "sebut", "daftar", "definisi", "identifikasi"],
    "Understand": ["jelaskan", "mengapa", "bagaimana", "ringkas", "uraikan", "makna"],
    "Apply": ["gunakan", "menerapkan", "contoh penerapan", "terapkan", "praktikkan", "implementasikan"],
}


def validate_bloom_question(question, level):
    if level not in bloom_keywords:
        return False
    q = question.lower()
    return any(k in q for k in bloom_keywords[level])

# ----------------------------------------------------------
# 5. Prompt Generator (Pertanyaan Saja)
# ----------------------------------------------------------
SYSTEM_PROMPT = r"""
Anda adalah generator pertanyaan Taksonomi Bloom khusus level 1–3:
- Remember (C1)
- Understand (C2)
- Apply (C3)

============================================================
BATASAN FOKUS LEVEL
============================================================
- Hanya buat pertanyaan Bloom level 1–3.
- Dilarang membuat pertanyaan Analyze, Evaluate, atau Create.
- Pastikan seluruh pertanyaan dapat dijawab langsung dari konteks.

============================================================
ATURAN KETAT
============================================================
1. Pertanyaan HARUS menggunakan istilah spesifik yang muncul dalam konteks.
2. Dilarang menambah informasi baru.
3. Dilarang membuat asumsi atau inferensi tambahan.
4. Pertanyaan harus berdiri sendiri (standalone) dan jelas tanpa membaca konteks.
5. Dilarang menggunakan kata ganti seperti "hal tersebut", "peristiwa ini", dll.

============================================================
TARGET OUTPUT
============================================================
- Hasilkan 8 sampai 15 pertanyaan valid per konteks.
- Semua pertanyaan harus berasal dari tiga level Bloom berikut:
  * Remember (minimal 3 pertanyaan)
  * Understand (minimal 3 pertanyaan)
  * Apply (minimal 2 pertanyaan, bila konteks memungkinkan)
- Jika konteks tidak mendukung Apply, hasilkan lebih banyak Remember dan Understand.

============================================================
FORMAT OUTPUT
============================================================
Keluarkan JSON array saja:

[
  {
    "bloom_level": "...",
    "question": "...",
    "context_list": [...]
  }
]
"""

def generate_bloom_items(concept_ids, concept_texts):
    context_joined = "\n".join([f"[ctx:{cid}] {txt}" for cid, txt in concept_texts.items()])
    ctx_list_str = json.dumps(concept_ids)

    USER_PROMPT = f"""
Konteks:
{context_joined}

Gunakan context_list persis seperti ini:
{ctx_list_str}

Instruksi:
- Hasilkan 8–15 pertanyaan Bloom level 1–3.
- Level yang diizinkan: Remember, Understand, Apply.
- Jangan buat Analyze/Evaluate/Create.
- Semua pertanyaan harus dapat dijawab langsung dari konteks.
- Gunakan istilah yang persis muncul dalam konteks.
- Dilarang menambah informasi baru.
- Pertanyaan harus standalone dan tidak boleh menggunakan frasa ambigu.

PENTING:
- Semua pertanyaan harus bersifat mandiri (standalone).
- Setiap pertanyaan harus menyebut ulang minimal satu istilah inti dari konteks,
  seperti nama tokoh, objek, peristiwa, konsep, langkah, atau istilah teknis
  yang muncul dalam konteks.
- Dilarang menggunakan kata rujukan ambigu seperti:
  "kisah tersebut", "kejadian ini", "solusi itu", "para ahli yang terlibat",
  "peristiwa ini", "hal tersebut", "kondisi ini", atau kata ganti rujukan lain.
- Pertanyaan harus tetap jelas, lengkap, dan dapat dipahami tanpa membaca konteks.
- Dilarang membuat pertanyaan yang hanya menggunakan kata ganti,
  atau tidak menyebut ulang istilah dalam konteks.
- Jika konteks membahas objek tertentu (misalnya "truk yang tersangkut di terowongan"),
  maka setiap pertanyaan HARUS menyebut objek tersebut secara eksplisit.
- Jika konteks membahas teori atau istilah tertentu, maka pertanyaan HARUS menyebut kembali
  nama teori atau istilah tersebut.
- Jika pertanyaan terkait studi kasus dalam konteks, pastikan pertanyaan menyebutkan nama atau detail studi kasus tersebut.
- Dilarang membuat pertanyaan yang hanya menggunakan istilah generik seperti
  "prototipe", "desainer", "konsep tersebut", "proses itu", tanpa merujuk pada istilah spesifik
  yang muncul dalam konteks.

- Jika konteks menyebut detail seperti "menggunakan prototipe kertas berfidelitas rendah"
  maka pertanyaan TIDAK BOLEH hanya berbentuk:
    "Apa kekurangan prototipe dalam konteks?"
  tetapi harus berbentuk:
    "Apa kekurangan prototipe kertas berfidelitas rendah yang digunakan dalam proses desain?"
  
- Setiap pertanyaan HARUS menyebutkan kembali frasa inti dari konteks,
  bukan versi ringkas atau generiknya.
- Tidak boleh membuat pertanyaan dengan "yang disebutkan dalam konteks" atau variasinya. Harus menyebut kembali konteks intinya dalam pertanyaan.

PENTING:
- Dilarang menyebut label konteks seperti "[ctx:75]" atau "[ctx:111]" dalam pertanyaan.
- Pertanyaan harus menyebut kembali istilah inti dari teks, BUKAN ID konteks.
- Dilarang membuat pertanyaan dengan frasa "dalam konteks [ctx:...]" atau variasinya.


Format output:
Keluarkan hanya JSON array, tanpa teks lain, dengan format setiap elemen seperti:
{{
  "bloom_level": "...",
  "question": "...",
  "context_list": [...]
}}

Tidak boleh ada penjelasan atau teks apa pun di luar JSON array.
"""
    log_var("USER_PROMPT", USER_PROMPT, preview=400)
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": USER_PROMPT},
            ],
            temperature=0.5,
            max_tokens=2000
        )

        content = response.choices[0].message.content.strip()
        content = content.replace("```json", "").replace("```", "").strip()

        items = robust_json_parse(content)
        log_var("Generated Items", items, preview=400)

        validated = []
        for item in items:
            if "bloom_level" not in item or "question" not in item:
                continue

            lvl = item["bloom_level"]
            # ✅ DIPERBAIKI: Validasi lebih aman
            if validate_bloom_question(item["question"], lvl):
                # Pastikan context_list konsisten
                item["context_list"] = concept_ids
                validated.append(item)

        return validated, context_joined

    except Exception as e:
        logger.info(f"❌ Error generate_bloom_items: {e}")
        return [], context_joined

# ----------------------------------------------------------
# 6. JSON Parser
# ----------------------------------------------------------
def robust_json_parse(text):
    text = text.strip()
    candidates = re.findall(r"\[[\s\S]*\]", text)
    for cand in candidates:
        try:
            obj = json.loads(cand)
            if isinstance(obj, list):
                return obj
        except:
            pass
    return []

# ----------------------------------------------------------
# 7. Main Generate Loop - DIPERBAIKI
# ----------------------------------------------------------
results = []
failed_concepts = []

jsonl_path = f"{output_dir}/bloom_1_3_{timestamp}.jsonl"
jsonl_file = open(jsonl_path, "w", encoding="utf-8")

logger.info("🚀 Memulai generasi pertanyaan...")

for i, row in tqdm(df.iterrows(), total=total_rows, desc="Generating Questions"):
    cid = row["chunk_id"]
    text = row["chunk_text"]

    try:
        #selected_ids = pick_related_concepts(cid, text, df, top_k=2)
        #selected_ids = pick_related_concepts_adaptive(cid, text, df)
        selected_ids = [cid]   # hanya pakai satu konteks


        # if len(selected_ids) < 2:
        #     logger.info(f"⏭️ Skip {cid}: Tidak ada konsep terkait")
        #     continue

        # ✅ DIPERBAIKI: Cara akses DataFrame yang benar
        # selected_texts = {
        #     sid: df.loc[df["id"] == sid, "context"].values[0]
        #     for sid in selected_ids
        # }
        selected_texts = {cid: text}


        items, ctx_joined = generate_bloom_items(selected_ids, selected_texts)

        if not items:
            logger.info(f"⚠️ Tidak ada pertanyaan valid untuk {cid}")
            failed_concepts.append(cid)
            continue

        for item in items:
            rec = {
                "concept_id_primary": cid,
                "context_list": item["context_list"],
                "context_text": ctx_joined,
                "bloom_level": item["bloom_level"],
                "question": item["question"],
                "timestamp": timestamp
            }
            results.append(rec)
            jsonl_file.write(json.dumps(rec, ensure_ascii=False) + "\n")

    except Exception as e:
        logger.info(f"❌ Error processing {cid}: {e}")
        failed_concepts.append(cid)
        continue

jsonl_file.close()

# ----------------------------------------------------------
# 8. Simpan Excel & Format
# ----------------------------------------------------------
if results:
    df_out = pd.DataFrame(results)
    xlsx_path = f"{output_dir}/bloom_1_3_{timestamp}.xlsx"
    df_out.to_excel(xlsx_path, index=False)
    
    # Format Excel
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active

        font = Font(name="Calibri", size=11)
        align = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = align

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 55)

        wb.save(xlsx_path)
        logger.info(f"✅ Excel formatted: {xlsx_path}")
        
    except Exception as e:
        logger.info(f"⚠️ Gagal format Excel: {e}")

# ----------------------------------------------------------
# 9. Final Report
# ----------------------------------------------------------
logger.info(f"\n📊 LAPORAN AKHIR:")
logger.info(f"✅ Konsep berhasil: {len(results)} pertanyaan dari {total_rows - len(failed_concepts)}/{total_rows} konsep")
logger.info(f"❌ Konsep gagal: {len(failed_concepts)}")

if results:
    final_df = pd.DataFrame(results)
    logger.info(f"🎯 Distribusi Level Bloom:")
    logger.info(final_df['bloom_level'].value_counts())

logger.info(f"\n🎉 GENERASI PERTANYAAN SELESAI!")
logger.info(f"📁 Output: {output_dir}")
logger.info(f"📄 JSONL: {jsonl_path}")