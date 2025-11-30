# ==========================================================
# Generate Natural Answers + Evidence + Source IDs (AVILA Version)
# ==========================================================

import os
import json
import re
import pandas as pd
from tqdm import tqdm
from datetime import datetime
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from dotenv import load_dotenv
import sys
import os
# memastikan folder project menjadi root import
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)
from utils_common.utils import *
# ----------------------------------------------------------
# 1. Konfigurasi
# ----------------------------------------------------------
load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#INPUT_JSONL = "datasets/merged_dataset20251130_074344.jsonl"
INPUT_JSONL = "datasets/answers_20251130_074721-tambahan.jsonl"

#versi XLSX
# INPUT_XLSX = "datasets/answers_20251130_tambahan.xlsx"
# df_questions = pd.read_excel(INPUT_XLSX)
# questions = df_questions.to_dict(orient="records")


OUTPUT_DIR = f"results/answers/answers_{timestamp}"
os.makedirs(OUTPUT_DIR, exist_ok=True)

output_jsonl_path = f"{OUTPUT_DIR}/answers_{timestamp}.jsonl"
import logging
setup_logging(OUTPUT_DIR)
logger = logging.getLogger(__name__)
import json
import re
import time


def split_contexts(context_text):
    """
    Memecah context_text_joined menjadi mapping {ctx_id: text}
    """
    blocks = re.split(r'\[ctx:(\d+)\]', context_text)
    result = {}
    for i in range(1, len(blocks), 2):
        cid = blocks[i]
        text = blocks[i+1].strip()
        result[cid] = text
    return result

def safe_json_extract(text):
    """
    Ekstraktor JSON robust:
    - mencari {...}
    - membersihkan karakter ilegal
    - fallback jika gagal
    """

    # bersihkan markdown
    text = text.replace("```json", "").replace("```", "").strip()

    # cari blok JSON
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        json_str = match.group(0)

        # perbaiki kutip ganda yang rusak
        json_str = json_str.replace("\n", " ").replace("\r", " ")

        # hilangkan karakter NUL atau non UTF-8
        json_str = re.sub(r"[\x00-\x1F\x7F]", "", json_str)

        try:
            return json.loads(json_str)
        except:
            pass

    # fallback jika tetap gagal
    return None

# ----------------------------------------------------------
# 2. Prompt Sistem
# ----------------------------------------------------------
SYSTEM_PROMPT = """
Anda adalah generator jawaban berbasis konteks. Ikuti semua aturan berikut secara ketat agar jawaban memiliki kualitas tinggi dan konsisten dengan evaluasi RAGAS.

============================================================
PRINSIP FUNDAMENTAL (WAJIB)
============================================================
1. Semua jawaban HARUS sepenuhnya berasal dari konteks.
2. Tidak boleh menambah, menebak, mengeneralisasi, menyimpulkan,
   atau menginterpolasi informasi yang tidak tertulis di konteks.
3. Jawaban harus akurat, spesifik, dan terkait langsung dengan isi konteks.
4. Jawaban tidak boleh bersifat umum, abstrak, atau mengandung opini.
5. Jika konteks terbatas atau tidak lengkap:
   - berikan jawaban paling dekat yang secara eksplisit muncul dalam konteks
   - tanpa membuat tambahan pengetahuan baru.

============================================================
ATURAN PENTING UNTUK RAGAS
============================================================
Untuk memaksimalkan skor RAGAS:

A. ContextPrecision:
   - Gunakan HANYA bagian konteks yang benar benar diperlukan.
   - Jangan menggunakan detail yang tidak relevan.

B. ContextRecall:
   - Jika jawaban membutuhkan lebih dari satu bagian konteks,
     sertakan semua elemen relevan dari konteks tersebut.

C. Faithfulness:
   - Jawaban harus mengikuti kalimat dan makna asli konteks.
   - Parafrase hanya boleh dilakukan bila makna tetap identik.

D. AnswerRelevancy:
   - Jawaban harus secara langsung menjawab pertanyaan.
   - Hindari jawaban terlalu pendek atau terlalu umum.

E. AnswerCorrectness:
   - Pastikan jawaban mereproduksi fakta dari konteks secara tepat.

============================================================
ATURAN UNTUK JAWABAN DEFINISI
============================================================
Jika pertanyaan dimulai dengan:
- "Apa yang dimaksud ..."
- "Apa itu ..."
- "Apa definisi ..."

MAKA:
1. Jawaban WAJIB dimulai dengan:
   "<Istilah> adalah ..." atau "<Istilah> merupakan ..."
2. Istilah harus diambil persis seperti dalam pertanyaan.
3. Definisi harus persis berdasarkan isi konteks:
   - bisa berupa kutipan langsung
   - atau parafrase yang sangat dekat
4. Tidak boleh menambah contoh, analogi, atau generalisasi.

============================================================
ATURAN EVIDENCE
============================================================
1. Evidence HARUS berupa kutipan langsung dari konteks,
   menggunakan teks persis seperti sumbernya.
2. Evidence harus merupakan bagian yang digunakan secara nyata
   untuk membentuk jawaban.
3. Evidence tidak boleh diparafrasa.
4. answer_source_ids HARUS berisi ID konteks yang benar benar dipakai.

============================================================
OUTPUT WAJIB DALAM FORMAT BERIKUT
============================================================
{
  "answer": "<jawaban berbasis konteks>",
  "answer_source_ids": ["<id1>", "<id2>"],
  "evidence": "<teks asli dari konteks yang digunakan>"
}

Tidak boleh ada teks lain di luar JSON.
"""
# ----------------------------------------------------------
# 3. Fungsi Generate Jawaban + Evidence
# ----------------------------------------------------------
def generate_answer_with_evidence(question, context_text, context_ids):
    USER_PROMPT = f"""
Konteks:
{context_text}

Pertanyaan:
{question}

Kembalikan output dalam format JSON berikut dan TIDAK BOLEH ada teks lain:
{{
  "answer": "...",
  "answer_source_ids": [...],
  "evidence": "...",
  "chunk_groundtruth":[...]
}}
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": USER_PROMPT}
            ],
            temperature=0.2,
            max_tokens=700
        )

        raw = response.choices[0].message.content.strip()

        data = safe_json_extract(raw)

        # jika JSON tidak bisa diparse → fallback
        if data is None or not isinstance(data, dict):
            return {
                "answer": "Konteks tidak menyediakan informasi untuk menjawab pertanyaan ini.",
                "answer_source_ids": [],
                "evidence": ""
            }

        # validasi minimal field
        if not all(k in data for k in ["answer", "answer_source_ids", "evidence"]):
            return {
                "answer": "Konteks tidak menyediakan informasi untuk menjawab pertanyaan ini.",
                "answer_source_ids": [],
                "evidence": ""
            }

        return data

    except Exception as e:
        logger.info(f"❌ Exception: {e}")
        return {
            "answer": "Konteks tidak menyediakan informasi untuk menjawab pertanyaan ini.",
            "answer_source_ids": [],
            "evidence": ""
        }

# ----------------------------------------------------------
# 4. Muat Data Pertanyaan
# ----------------------------------------------------------
questions = []
with open(INPUT_JSONL, "r", encoding="utf-8") as f:
    for line in f:
        questions.append(json.loads(line))

# batasi 10 dulu atas dan bawah
#top5 = questions[:5]
#bottom5 = questions[-5:]

# Jika total data kurang dari 10, top5 dan bottom5 bisa overlap
# maka kita buang duplikasi
#questions = top5 + bottom5
questions = list({id(q): q for q in questions}.values())  # dedupe by object id


logger.info(f"📘 Loaded {len(questions)} questions")

# ----------------------------------------------------------
# 5. Generate Answers
# ----------------------------------------------------------
output_jsonl = open(output_jsonl_path, "w", encoding="utf-8")
results = []

logger.info("🚀 Generating answers with evidence...")

for item in tqdm(questions, desc="Answering Questions"):
    time.sleep(0.15)

    q = item["question"]
    ctx = item["context_text"]
    ctx_ids = item["context_list"]

    result = generate_answer_with_evidence(q, ctx, ctx_ids)
    chunk_map = split_contexts(ctx)
    used_chunks = {}

    for cid in result.get("answer_source_ids", []):
        cid = re.sub(r"[^0-9]", "", str(cid))
        if cid in chunk_map:
            used_chunks[cid] = chunk_map[cid]
    # NEW: hanya ID saja
    used_chunk_ids = list(used_chunks.keys())

    rec = {
        "concept_id_primary": item["concept_id_primary"],
        "context_list": ctx_ids,
        "context_groundtruth": ctx,
        "bloom_level": item["bloom_level"],
        "question": q,
        "answer_groundtruth": result.get("answer", ""),
        "answer_source_ids": result.get("answer_source_ids", []),
        "evidence": result.get("evidence", ""),
        "used_context": used_chunks,
        "used_context_ids": used_chunk_ids,
        "timestamp": timestamp
    }



    results.append(rec)
    output_jsonl.write(json.dumps(rec, ensure_ascii=False) + "\n")

output_jsonl.close()

# ----------------------------------------------------------
# 6. Simpan Excel
# ----------------------------------------------------------
df_out = pd.DataFrame(results)
xlsx_path = f"{OUTPUT_DIR}/answers_{timestamp}.xlsx"
df_out.to_excel(xlsx_path, index=False)

# Format kolom supaya rapi
try:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.freeze_panes = "A2"

    font = Font(name="Calibri", size=11)
    align = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = align

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 55)

    wb.save(xlsx_path)
    logger.info(f"✅ Excel formatted: {xlsx_path}")

except Exception as e:
    logger.info(f"⚠️ Failed formatting Excel: {e}")

# ----------------------------------------------------------
# 7. Final Report
# ----------------------------------------------------------
logger.info("\n🎉 SELESAI GENERATE JAWABAN + EVIDENCE")
logger.info(f"📁 Output Folder: {OUTPUT_DIR}")
logger.info(f"📄 JSONL: {output_jsonl_path}")
logger.info(f"📊 Total jawaban: {len(results)}")
