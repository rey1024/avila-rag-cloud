import os
import logging
import json
import yaml
import faiss
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from sentence_transformers import SentenceTransformer, util
import torch

from datasets import Dataset
from langchain_openai import ChatOpenAI, OpenAIEmbeddings

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

import requests
import openpyxl
global_results_generation_path = "results/generation"

def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def fmt(v, d=2):
    try:
        return round(float(v), d)
    except:
        return 0.0


def setup_logging(exp, timestamp):
    log_dir = f"{global_results_generation_path}/Gen{timestamp}/{exp['id']}_{timestamp}"
    ensure_dir(log_dir)
    log_file = f"{log_dir}/log_{timestamp}.log"
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # remove handlers lama (jika ada)
    if logger.hasHandlers():
        logger.handlers.clear()

    # format log
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # 1) Handler untuk FILE
    fh = logging.FileHandler(log_file, mode="w", encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    # 2) Handler untuk TERMINAL
    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(formatter)
    logger.addHandler(sh)


def log_var(name, v, preview=200):
    """Log a compact summary of a variable: type, size/shape and a small preview."""
    logger = logging.getLogger(__name__)
    try:
        t = type(v).__name__
        size = None
        preview_val = None

        # numpy / torch
        if hasattr(v, "shape"):
            size = getattr(v, "shape")
        elif isinstance(v, (list, tuple, dict)):
            size = len(v)
        elif isinstance(v, str):
            size = len(v)

        # helper to keep only first N words
        def first_n_words(s, n=5):
            try:
                words = str(s).split()
                if len(words) > n:
                    return " ".join(words[:n]) + "..."
                return " ".join(words)
            except Exception:
                return str(s)[:preview]

        # preview
        if isinstance(v, str):
            preview_val = first_n_words(v, 5)
        elif isinstance(v, dict):
            try:
                # show up to 3 key:shortvalue pairs
                items = list(v.items())[:3]
                preview_val = {k: first_n_words(vv, 5) for k, vv in items}
            except Exception:
                preview_val = str(list(v.keys())[:3])
        elif isinstance(v, (list, tuple)):
            # if list of strings, trim each element to first 5 words
            try:
                sample = v[:3]
                preview_list = []
                for el in sample:
                    if isinstance(el, str):
                        preview_list.append(first_n_words(el, 5))
                    else:
                        preview_list.append(el)
                preview_val = preview_list
            except Exception:
                preview_val = str(v)[:preview]
        else:
            # try numpy conversion
            try:
                arr = np.asarray(v)
                preview_flat = arr.flatten()
                preview_val = preview_flat[:3].tolist()
            except Exception:
                preview_val = str(v)[:preview]

        logger.info(f"    - {name}: type={t}, size={size}, preview={preview_val}")
    except Exception as e:
        logger.info(f"    - {name}: (error summarizing: {e})")

def get_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def load_eval_dataset(path):
    df = pd.read_excel(path)
    
   # Ambil 5 teratas dan 5 terbawah
#

    #df = df[df["id"] == 2].copy()
    #df =  df.head(5)

    df.columns = df.columns.str.lower()

    #ambil sample masing2 dari bloomlevel
    
    required = ["id", "context", "question", "answer_groundtruth", "bloom_level"]
    for c in required:
        if c not in df:
            raise ValueError(f"Kolom {c} tidak ada di dataset evaluasi")

    #Ambil 5 item per bloomlevel
    # df = (
    #     df.groupby("bloomlevel")
    #     .head(2)
    #     .reset_index(drop=True)
    # )





    data = []
    for _, r in df.iterrows():
        data.append({
            "id": r["id"],
            "context": r["context"],
            "question": r["question"],
            "answer_groundtruth": r["answer_groundtruth"],
            "bloom_level": r["bloom_level"]
        })


    print(f"Dataset evaluasi: {len(data)} item")
    return data, df

def to_numpy(x):
    if hasattr(x, "cpu"):
        return x.cpu().numpy()
    return np.asarray(x)


def save_excel_report(
    exp_id, out_dir, timestamp,
    df_source,
    questions,
    predictions,
    references,
    contexts_used,
    ragas_scores_item,
    ragas_agg,
    chunks,
    chunk_ids,
    retrieval_hits,
    semantic_scores,
    embedder,
    gt_chunk_ids,
    gt_chunk_retrieved,
    gt_chunk_rank,
    gt_chunk_rr,
    gt_topk_sims,
    faiss_queries,
    existing_chunk_embs

):


    excel_path = f"{out_dir}/{exp_id}_{timestamp}.xlsx"


    df_items = df_source.copy()

    chunk_to_id = {chunks[i]: chunk_ids[i] for i in range(len(chunks))}


    # -------------------------
    # Format contexts
    # -------------------------
    formatted = []
    retrieved_ids_list = []
    for retrieved in contexts_used:
        ids = [chunk_to_id[c] for c in retrieved]
        retrieved_ids_list.append(ids)
        combined = "\n".join([f"[{cid}] {txt}" for cid, txt in zip(ids, retrieved)])
        formatted.append(combined)
    df_items["faiss_query"] = faiss_queries

    df_items["contexts_retrieval"] = formatted
    df_items["retrieved_chunk_ids"] = retrieved_ids_list
    df_items["retrieved_similarity"] = semantic_scores

    # -------------------------
    # Answers
    # -------------------------
    df_items["llm_answer"] = predictions

    df_items["context_precision"] = [x["context_precision"] for x in ragas_scores_item]
    df_items["context_recall"] = [x["context_recall"] for x in ragas_scores_item]
    df_items["answer_relevancy"] = [x["answer_relevancy"] for x in ragas_scores_item]
    df_items["faithfulness"] = [x["faithfulness"] for x in ragas_scores_item]
#    df_items["answer_correctness"] = [x["answer_correctness"] for x in ragas_scores_item]

    df_items["retrieval_hit"] = retrieval_hits
    df_items["gt_chunk_id"] = gt_chunk_ids
    df_items["gt_chunk_retrieved"] = gt_chunk_retrieved
    df_items["gt_chunk_rank"] = gt_chunk_rank
    df_items["gt_chunk_rr"] = gt_chunk_rr

    # -------------------------
    # Summary
    # -------------------------
    df_sum = pd.DataFrame({
        "metric": list(ragas_agg.keys()),
        "score": list(ragas_agg.values())
    })

    # -------------------------
    # CHUNKS SHEET
    # -------------------------
    #chunk_embs = to_numpy(embedder.encode(chunks, normalize_embeddings=True))
    chunk_embs = existing_chunk_embs   # dari run_generation()

    #context_embs = embedder.encode(df_source["context"].tolist(), normalize_embeddings=True)
    context_embs = embedder.encode(
        df_source["context"].tolist(),
        normalize_embeddings=True,
        convert_to_numpy=True
    )

    #
    
    sim_mat = util.cos_sim(
        torch.tensor(chunk_embs),
        torch.tensor(context_embs)
    ).cpu().numpy()

    df_chunks = pd.DataFrame({
        "chunk_id": range(1, len(chunks)+1),
        "chunk_text": chunks,
        "chars": [len(c) for c in chunks],
        "tokens": [len(c.split()) for c in chunks],
        "top_groundtruth_similarity": sim_mat.max(axis=1)
    })

    # ======================================================================
    # Tambahan Visualisasi untuk Excel
    # ======================================================================

    import matplotlib.pyplot as plt
    import seaborn as sns
    from openpyxl.drawing.image import Image as XLImage
    import numpy as np
    import io

    # ------------------------------------------------------------
    # STEP 1: assign bloom level dominan untuk setiap chunk
    # ------------------------------------------------------------
    nearest_bloom = []
    for i in range(sim_mat.shape[0]):
        j = np.argmax(sim_mat[i])  # context yang paling mirip
        nearest_bloom.append(df_source["bloom_level"].iloc[j])

    df_chunks["nearest_bloom"] = nearest_bloom
    # ============================================================
    # SET BLOOM ORDER
    # ============================================================
    bloom_order = [
        "1. Remember",
        "2. Understand",
        "3. Apply",
        "4. Analyze",
        "5. Evaluate",
        "6. Create",
    ]

    df_chunks["nearest_bloom"] = pd.Categorical(
        df_chunks["nearest_bloom"],
        categories=bloom_order,
        ordered=True
    )

    # ------------------------------------------------------------
    # STEP 2: Buat folder grafik
    # ------------------------------------------------------------
    plot_dir = f"{out_dir}/plots"
    os.makedirs(plot_dir, exist_ok=True)

    # ------------------------------------------------------------
    # STEP 3-A: BOX PLOT
    # ------------------------------------------------------------
    plt.figure(figsize=(10,6))
    sns.boxplot(
        data=df_chunks,
        x="nearest_bloom",
        y="top_groundtruth_similarity"
    )
    plt.xlabel("Bloom Level (nearest groundtruth)")
    plt.ylabel("Top Groundtruth Similarity")
    plt.title("Boxplot Top Groundtruth Similarity per Bloom Level")
    plt.tight_layout()

    boxplot_path = f"{plot_dir}/boxplot_similarity_per_bloom.png"
    plt.savefig(boxplot_path, dpi=300)
    plt.close()


    # ------------------------------------------------------------
    # STEP 3-B: VIOLIN PLOT
    # ------------------------------------------------------------
    plt.figure(figsize=(10,6))
    sns.violinplot(
        data=df_chunks,
        x="nearest_bloom",
        y="top_groundtruth_similarity",
        inner="quartile"
    )
    plt.xlabel("Bloom Level (nearest groundtruth)")
    plt.ylabel("Top Groundtruth Similarity")
    plt.title("Violin Plot Similarity per Bloom Level")
    plt.tight_layout()

    violin_path = f"{plot_dir}/violin_similarity_per_bloom.png"
    plt.savefig(violin_path, dpi=300)
    plt.close()


    # ------------------------------------------------------------
    # STEP 3-C: HEATMAP RINGKAS
    # (median similarity tiap Bloom dikali chunk range)
    # ------------------------------------------------------------
    pivot = df_chunks.pivot_table(
        index="nearest_bloom",
        values="top_groundtruth_similarity",
        aggfunc="median"
    ).sort_index()

    plt.figure(figsize=(6,4))
    sns.heatmap(
        pivot.T,
        annot=True,
        cmap="viridis",
        fmt=".2f"
    )
    plt.title("Median Similarity per Bloom Level")
    plt.tight_layout()

    heatmap_path = f"{plot_dir}/heatmap_similarity_per_bloom.png"
    plt.savefig(heatmap_path, dpi=300)
    plt.close()


    # ------------------------------------------------------------
    # STEP 4: summary statistik ke DataFrame
    # ------------------------------------------------------------
    df_bloom_summary = df_chunks.groupby("nearest_bloom")["top_groundtruth_similarity"].describe()
    df_bloom_summary.reset_index(inplace=True)

    # -------------------------
    # Write
    # -------------------------
    with pd.ExcelWriter(excel_path, engine="openpyxl") as w:
        df_sum.to_excel(w, index=False, sheet_name="summary")
        df_items.to_excel(w, index=False, sheet_name="results")
        df_chunks.to_excel(w, index=False, sheet_name="chunks")

        # sheet baru: Bloom Summary
        df_bloom_summary.to_excel(w, index=False, sheet_name="bloom_summary")

        # sheet baru: Visualizations
        ws = w.book.create_sheet("visualizations")

        # masukkan gambar boxplot
        img1 = XLImage(boxplot_path)
        ws.add_image(img1, "A1")

        # masukkan gambar violinplot
        img2 = XLImage(violin_path)
        ws.add_image(img2, "A30")

        # masukkan heatmap
        img3 = XLImage(heatmap_path)
        ws.add_image(img3, "A60")
        # ----------------------------------------
        # APPLY FORMATTING (wrap, align, auto width)
        # ----------------------------------------
        wb = w.book
        from openpyxl.styles import Alignment

        for sheet_name in ["results", "summary", "chunks"]:
            ws = wb[sheet_name]
                # === Freeze Top Row ===
            ws.freeze_panes = "A2"
            # wrap + alignment
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                        horizontal="left"
                    )

            # auto width per column
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_length:
                        max_length = len(val)

                ws.column_dimensions[col_letter].width = min(max_length + 3, 80)



    print(f"✔ Excel tersimpan: {excel_path}")
    return excel_path

def format_excel(xlsx_path):
    """
    Menerapkan formatting seperti save_excel_report()
    - Freeze baris pertama
    - Wrap text
    - Vertical top align
    - Auto column width (max 80)
    """

    wb = load_workbook(xlsx_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Freeze baris pertama
        ws.freeze_panes = "A2"

        # Wrap text + alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal="left"
                )

        # Auto width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_length:
                    max_length = len(v)

            ws.column_dimensions[col_letter].width = min(max_length + 3, 80)

    wb.save(xlsx_path)
    print(f"✔ Format Excel selesai: {xlsx_path}")

def load_pdf_texts(pdf_dir):
    import pypdf
    from docx import Document

    base = Path(pdf_dir)
    pdfs = list(base.glob("*.pdf"))
    docs = list(base.glob("*.docx"))

    if not pdfs and not docs:
        raise ValueError(f"Tidak ada file PDF/DOCX di {pdf_dir}")

    texts = []

    for p in pdfs:
        reader = pypdf.PdfReader(str(p))
        t = " ".join(page.extract_text() or "" for page in reader.pages)
        texts.append(t)

    for d in docs:
        doc = Document(str(d))
        lines = [p.text for p in doc.paragraphs if p.text]
        texts.append("\n".join(lines))

    return "\n".join(texts)

def auto_output_folder(exp_id, timestamp):
    folder = f"{global_results_generation_path}/Gen{timestamp}/{exp_id}_{timestamp}"
    ensure_dir(folder)
    return folder


def save_global_summary(summary_dict, timestamp):
    """
    Membuat 1 Excel summary keseluruhan eksperimen.
    summary_dict adalah dictionary berikut:

    summary[exp_id] = {
        "status": "ok",
        "output_dir": "...",
        "excel": "....xlsx",
        "scores": { ... ragas_agg ... }
    }
    """
    rows = []

    for exp_id, info in summary_dict.items():
        scores = info["scores"]

        rows.append({
            "experiment": exp_id,
            "context_precision": scores.get("context_precision", 0),
            "context_recall": scores.get("context_recall", 0),
            "answer_relevancy": scores.get("answer_relevancy", 0),
            "faithfulness": scores.get("faithfulness", 0),
            #"answer_correctness": scores.get("answer_correctness", 0),
            "output_dir": info["output_dir"],
            "excel_file": info["excel"]
        })

    df = pd.DataFrame(rows)

    out_path = f"{global_results_generation_path}/Gen{timestamp}/summary_all_{timestamp}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="summary_all")

        # formatting
        ws = w.book["summary_all"]
        from openpyxl.styles import Alignment

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal="left"
                )

        # auto width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[col_letter].width = min(max_length + 3, 80)

    print(f"✔ Global summary Excel tersimpan: {out_path}")
    return out_path
