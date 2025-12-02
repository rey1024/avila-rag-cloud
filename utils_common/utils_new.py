import os
import logging
from datetime import datetime
import numpy as np

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def get_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def setup_logging(log_dir):
    """Setup logging untuk file dan terminal."""
    ensure_dir(log_dir)

    log_file = f"{log_dir}/log_{get_timestamp()}.log"

    # root logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # bersihkan handler lama
    for h in logger.handlers[:]:
        logger.removeHandler(h)

    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # handler FILE
    fh = logging.FileHandler(log_file, mode="w", encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    # handler TERMINAL
    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(formatter)
    logger.addHandler(sh)

    logger.info(f"Log file created at: {log_file}")
    return log_file


class HTMLFormatter(logging.Formatter):
    """Custom formatter for HTML logging with colors and collapsible sections."""

    def __init__(self):
        super().__init__()
        self.color_map = {
            'DEBUG': 'blue',
            'INFO': 'green',
            'WARNING': 'orange',
            'ERROR': 'red',
            'CRITICAL': 'purple'
        }
        self.open_sections = 0  # Track open <details> tags

    def format(self, record):
        # Escape HTML special characters
        message = self.formatMessage(record)
        message = message.replace('&', '&amp;').replace('<', '<').replace('>', '>')

        level_color = self.color_map.get(record.levelname, 'black')

        # Check for section headers (messages starting with ===)
        if message.strip().startswith('===') and message.strip().endswith('==='):
            # Close previous section if open
            if self.open_sections > 0:
                html = f'</details>\n<h2 style="color: {level_color};">{message}</h2>\n<details open>\n<summary>{message}</summary>\n'
                self.open_sections += 1
            else:
                html = f'<h2 style="color: {level_color};">{message}</h2>\n<details open>\n<summary>{message}</summary>\n'
                self.open_sections += 1
        else:
            # Regular log message
            timestamp = datetime.fromtimestamp(record.created).strftime("%Y-%m-%d %H:%M:%S")
            html = f'<div style="margin-left: 20px; color: {level_color};">[{timestamp}] [{record.levelname}] {message}</div>\n'

        return html


def setup_html_logging(log_dir):
    """Setup logging untuk file, terminal, dan HTML."""
    ensure_dir(log_dir)

    log_file = f"{log_dir}/log_{get_timestamp()}.log"
    html_file = f"{log_dir}/log_{get_timestamp()}.html"

    # root logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # bersihkan handler lama
    for h in logger.handlers[:]:
        logger.removeHandler(h)

    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # handler FILE
    fh = logging.FileHandler(log_file, mode="w", encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    # handler TERMINAL
    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(formatter)
    logger.addHandler(sh)

    # Write HTML header
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write("""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Retrieval Log</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        details { margin-bottom: 10px; }
        summary { cursor: pointer; font-weight: bold; background-color: #f0f0f0; padding: 5px; border-radius: 3px; }
        h2 { margin-top: 30px; border-bottom: 1px solid #ccc; }
    </style>
</head>
<body>
""")

    # HTML file handler
    html_formatter = HTMLFormatter()
    hfh = logging.FileHandler(html_file, mode="a", encoding="utf-8")
    hfh.setLevel(logging.INFO)
    hfh.setFormatter(html_formatter)
    logger.addHandler(hfh)

    logger.info(f"Log file created at: {log_file}")
    logger.info(f"HTML log file created at: {html_file}")
    return log_file, html_file


def close_html_logging():
    """Close HTML logging by appending closing tags."""
    logger = logging.getLogger()
    for handler in logger.handlers:
        if isinstance(handler, logging.FileHandler) and handler.baseFilename.endswith('.html'):
            # Close any open sections and HTML
            with open(handler.baseFilename, 'a', encoding='utf-8') as f:
                # Close remaining open sections
                formatter = handler.formatter
                if hasattr(formatter, 'open_sections'):
                    for _ in range(formatter.open_sections):
                        f.write('</details>\n')
                f.write('</body>\n</html>\n')
            break


def log_var(name, v, preview=200):
    """Log a compact summary of a variable."""
    logger = logging.getLogger(__name__)
    try:
        t = type(v).__name__
        size = None
        preview_val = None

        if hasattr(v, "shape"):
            size = getattr(v, "shape")
        elif isinstance(v, (list, tuple, dict, str)):
            size = len(v)

        def first_n_words(s, n=5):
            words = str(s).split()
            return " ".join(words[:n]) + ("..." if len(words) > n else "")

        if isinstance(v, str):
            preview_val = first_n_words(v, 5)
        elif isinstance(v, dict):
            sample = list(v.items())[:3]
            preview_val = {k: first_n_words(val, 5) for k, val in sample}
        elif isinstance(v, (list, tuple)):
            sample = v[:3]
            preview_val = [
                first_n_words(el, 5) if isinstance(el, str) else el
                for el in sample
            ]
        else:
            try:
                arr = np.asarray(v)
                preview_val = arr.flatten()[:3].tolist()
            except:
                preview_val = str(v)[:preview]

        logger.info(f"    - {name}: type={t}, size={size}, preview={preview_val}")

    except Exception as e:
        logger.info(f"    - {name}: (error summarizing: {e})")


from openpyxl import load_workbook
from openpyxl.styles import Alignment

def format_excel(xlsx_path):
    """
    Menerapkan formatting seperti save_excel_report()
    - Freeze baris pertama
    - Wrap text
    - Vertical top align
    - Auto column width (max 80)
    """

    wb = load_workbook(xlsx_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Freeze baris pertama
        ws.freeze_panes = "A2"

        # Wrap text + alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal="left"
                )

        # Auto width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_length:
                    max_length = len(v)

            ws.column_dimensions[col_letter].width = min(max_length + 3, 80)

    wb.save(xlsx_path)
    print(f"✔ Format Excel selesai: {xlsx_path}")
