import os
import logging
from datetime import datetime
import numpy as np

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def get_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def setup_logging(log_dir):
    """Setup logging untuk file dan terminal."""
    ensure_dir(log_dir)

    log_file = f"{log_dir}/log_{get_timestamp()}.log"

    # root logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # bersihkan handler lama
    for h in logger.handlers[:]:
        logger.removeHandler(h)

    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # handler FILE
    fh = logging.FileHandler(log_file, mode="w", encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    # handler TERMINAL
    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(formatter)
    logger.addHandler(sh)

    logger.info(f"Log file created at: {log_file}")
    return log_file


def log_var(name, v, preview=200):
    """Log a compact summary of a variable."""
    logger = logging.getLogger(__name__)
    try:
        t = type(v).__name__
        size = None
        preview_val = None

        if hasattr(v, "shape"):
            size = getattr(v, "shape")
        elif isinstance(v, (list, tuple, dict, str)):
            size = len(v)

        def first_n_words(s, n=5):
            words = str(s).split()
            return " ".join(words[:n]) + ("..." if len(words) > n else "")

        if isinstance(v, str):
            preview_val = first_n_words(v, 5)
        elif isinstance(v, dict):
            sample = list(v.items())[:3]
            preview_val = {k: first_n_words(val, 5) for k, val in sample}
        elif isinstance(v, (list, tuple)):
            sample = v[:3]
            preview_val = [
                first_n_words(el, 5) if isinstance(el, str) else el
                for el in sample
            ]
        else:
            try:
                arr = np.asarray(v)
                preview_val = arr.flatten()[:3].tolist()
            except:
                preview_val = str(v)[:preview]

        logger.info(f"    - {name}: type={t}, size={size}, preview={preview_val}")

    except Exception as e:
        logger.info(f"    - {name}: (error summarizing: {e})")


from openpyxl import load_workbook
from openpyxl.styles import Alignment

def format_excel(xlsx_path):
    """
    Menerapkan formatting seperti save_excel_report()
    - Freeze baris pertama
    - Wrap text
    - Vertical top align
    - Auto column width (max 80)
    """

    wb = load_workbook(xlsx_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Freeze baris pertama
        ws.freeze_panes = "A2"

        # Wrap text + alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal="left"
                )

        # Auto width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_length:
                    max_length = len(v)

            ws.column_dimensions[col_letter].width = min(max_length + 3, 80)

    wb.save(xlsx_path)
    print(f"✔ Format Excel selesai: {xlsx_path}")
