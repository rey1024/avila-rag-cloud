
# run_retrieval_adv.py
import os, yaml, faiss, json
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

from pathlib import Path
from datetime import datetime
from sentence_transformers import SentenceTransformer
from nltk.tokenize import sent_tokenize
import nltk
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from rank_bm25 import BM25Okapi
import requests

nltk.download("punkt")
nltk.download("punkt_tab")

stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
import ast





# ============================================================
# UTIL
# ============================================================
def load_yaml(p):
    with open(p, "r", encoding="utf8") as f:
        return yaml.safe_load(f)

def ensure_dir(p):
    os.makedirs(p, exist_ok=True)

def auto_out(exp_id):
    d = f"outputsRetrievalADV/{stamp}/{exp_id}_{stamp}"
    ensure_dir(d)
    return d

# ============================================================
# PDF/DOCX
# ============================================================
def load_pdf_docx(folder):
    import pypdf
    from docx import Document
    texts = []
    p = Path(folder)
    for f in p.glob("*.pdf"):
        r = pypdf.PdfReader(str(f))
        texts.append(" ".join([pg.extract_text() or "" for pg in r.pages]))
    for d in p.glob("*.docx"):
        dc = Document(str(d))
        texts.append("\n".join([p.text for p in dc.paragraphs if p.text]))
    if not texts:
        raise ValueError("No PDF/DOCX found")
    return "\n".join(texts)

# ============================================================
# Chunking
# ============================================================
def fixed_chunk(text, size, overlap):
    t = text.split()
    step = size - overlap
    return [" ".join(t[i:i+size]) for i in range(0, len(t), step)]

def semantic_chunk_langchain(text, embedder, min_chunk_size=150, max_chunk_size=450, similarity_threshold=0.75):
    sents = sent_tokenize(text)
    if not sents:
        return []
    emb = embedder.encode(sents, normalize_embeddings=True, batch_size=4)
    chunks, cur, ln = [], [], 0
    for i, s in enumerate(sents):
        sl = len(s.split())
        if not cur:
            cur.append(s)
            ln += sl
            continue
        sim = np.dot(emb[i], emb[i-1])
        boundary = sim < similarity_threshold
        too_big = ln + sl > max_chunk_size
        if boundary or too_big:
            if ln < min_chunk_size and not too_big:
                cur.append(s)
                ln += sl
                continue
            chunks.append(" ".join(cur))
            cur, ln = [s], sl
        else:
            cur.append(s)
            ln += sl
    if cur:
        chunks.append(" ".join(cur))
    return chunks

def chunk_text(text, cfg, embedder):
    if cfg["type"] == "fixed":
        return fixed_chunk(text, cfg["size"], cfg["overlap"])
    if cfg["type"] == "semantic":
        return semantic_chunk_langchain(text, embedder, cfg.get("min_chunk_size",150), cfg.get("max_chunk_size",450), cfg.get("threshold",0.75))
    raise ValueError("Unknown chunking")

# ============================================================
# BM25
# ============================================================
def bm25_build(chunks):
    return BM25Okapi([c.split() for c in chunks])

def bm25_retrieve(model, q, k):
    scores = model.get_scores(q.split())
    ranked = np.argsort(scores)[::-1][:k]
    return ranked, scores

# ============================================================
# RRF
# ============================================================
def reciprocal_rank_fusion(ranks_list, max_k=60, k=60):
    scores = {}
    for lst in ranks_list:
        for pos, did in enumerate(lst):
            scores[did] = scores.get(did, 0) + 1/(k+pos+1)
    return [d for d,_ in sorted(scores.items(), key=lambda x:x[1], reverse=True)][:max_k]

# ============================================================
# Query Expansion
# ============================================================
def llm_rewrite_query(q):
    try:
        r = requests.post("http://localhost:11434/api/generate",
            json={"model":"gemma:7b-instruct","prompt":f"Rewrite:\n{q}","stream":False})
        return r.json().get("response",q).strip()
    except:
        return q

# ============================================================
# Evaluator Utils
# ============================================================
def token_overlap(a,b):
    sa, sb = set(a.split()), set(b.split())
    if not sa or not sb: return 0
    return len(sa & sb)/len(sa|sb)

def map_groundtruth_chunk(gt_text, chunks, chunk_embs, gt_emb):
    contain = [i for i,c in enumerate(chunks) if gt_text.strip() in c]
    if contain:
        return contain[0]
    overlaps = [token_overlap(gt_text,c) for c in chunks]
    ov_best = int(np.argmax(overlaps))
    if overlaps[ov_best] > 0.15:
        return ov_best
    sims = np.dot(chunk_embs, gt_emb)
    return int(np.argmax(sims))

def cluster_neighbors(idx, total, radius=1):
    return {i for i in range(idx-radius, idx+radius+1) if 0<=i<total}

def ndcg_at_k(retrieved, gt_id, k):
    gains = [1 if r==gt_id else 0 for r in retrieved[:k]]
    if not gains:
        return 0
    dcg = gains[0] + sum(g/np.log2(i+1) for i,g in enumerate(gains[1:], start=2))
    return dcg/1

def text_hit_match(gt_text, pairs, chunk_embs, gt_emb, overlap_threshold=0.25, semantic_threshold=0.86):
    be, bo, bs = 0,0,0
    for rid, ch in pairs:
        if gt_text.strip() in ch or ch in gt_text:
            be = 1
        ov = token_overlap(gt_text, ch)
        bo = max(bo, ov)
        sim = float(np.dot(chunk_embs[rid], gt_emb))
        bs = max(bs, sim)
    votes = (be==1) + (bo>=overlap_threshold) + (bs>=semantic_threshold)
    return int(votes>=2), be, bo, bs

def detect_misalignment(r):
    if r["text_hit@k"]==0:
        return "NOT_FOUND"
    if r["hit_at_k"]==1:
        return "ALIGNED"
    if r["cluster_hit@k"]==1:
        return "SHIFTED_NEAR"
    return "SHIFTED"

# ============================================================
# Visualization
# ============================================================
def plot_heatmap(mat, out, exp_id):
    plt.figure(figsize=(10,6))
    sns.heatmap(mat, cmap="viridis")
    p=f"{out}/heatmap.png"
    plt.title(exp_id)
    plt.tight_layout()
    plt.savefig(p); plt.close()
    return p

def plot_scatter_rank_similarity(df,out):
    plt.figure(figsize=(6,5))
    plt.scatter(df["rank"], df["topk_sim_max"], alpha=0.7)
    plt.xlabel("Rank"); plt.ylabel("TopK Sim"); plt.title("Rank vs Sim")
    p=f"{out}/scatter.png"
    plt.tight_layout(); plt.savefig(p); plt.close()
    return p

# ============================================================
# Excel formatting
# ============================================================
def format_decimal(ws):
    for r in ws.iter_rows():
        for c in r:
            try:
                float(c.value)
                c.number_format="0.00"
            except:
                pass

# ============================================================
# SAVE FAISS
# ============================================================
def save_faiss(out, index, embs, chunks):
    fd = f"{out}/faiss"; ensure_dir(fd)
    faiss.write_index(index, f"{fd}/index.faiss")
    np.save(f"{fd}/chunk_embs.npy", embs)
    with open(f"{fd}/chunks.jsonl","w",encoding="utf8") as f:
        for i,c in enumerate(chunks,1):
            f.write(json.dumps({"chunk_id":i,"text":c},ensure_ascii=False)+"\n")

# ============================================================
# ADVANCED RETRIEVAL + FULL EVALUATOR
# ============================================================
def run_retrieval_adv(exp, master, df):
    print(f"=== Running {exp['id']} ===")
    embedder = SentenceTransformer(exp["embedding"]["model"], device="cuda")
    corpus = load_pdf_docx(master["pdf_dir"])
    chunks = chunk_text(corpus, exp["chunking"], embedder)
    chunk_embs = embedder.encode(chunks, normalize_embeddings=True, batch_size=4)

    dim = chunk_embs.shape[1]
    index = faiss.IndexFlatL2(dim); index.add(chunk_embs)

    bm25 = bm25_build(chunks)
    top_k = master["top_k"]

    q_embs = embedder.encode(df["question"].tolist(), normalize_embeddings=True, batch_size=4)
    gt_embs = embedder.encode(df["context"].tolist(), normalize_embeddings=True, batch_size=4)

    full_sim_matrix = np.dot(chunk_embs, q_embs.T)

    rows=[]
    for i,row in df.iterrows():
        q = row["question"]
        if exp["retrieval"].get("query_expansion")=="llm_rewrite":
            q = llm_rewrite_query(q)
        q_emb = q_embs[i]
        gt_emb = gt_embs[i]
        gt_text = row["context"]

        # Retrieval
        _, idx_dense = index.search(q_emb.reshape(1,-1), top_k)
        faiss_ids = idx_dense[0]

        bm25_ids,_ = bm25_retrieve(bm25, q, top_k)

        if exp["retrieval"]["engine"]=="hybrid_rrf":
            retrieved = reciprocal_rank_fusion([faiss_ids,bm25_ids], max_k=top_k)
        elif exp["retrieval"]["engine"]=="bm25":
            retrieved = bm25_ids
        else:
            retrieved = faiss_ids

        # groundtruth chunk
        gt_chunk = map_groundtruth_chunk(gt_text, chunks, chunk_embs, gt_emb)

        # rank
        sims_q = np.dot(chunk_embs, q_emb)
        order = np.argsort(sims_q)[::-1]
        rank = int(np.where(order==gt_chunk)[0][0]) + 1
        rr = 1/rank

        # cluster
        cl = cluster_neighbors(gt_chunk, len(chunks), radius=1)
        cl_hit = int(any(r in cl for r in retrieved))

        # text hit
        pairs = [(rid, chunks[rid]) for rid in retrieved]
        text_hit, exact_sc, overlap_sc, sem_sc = text_hit_match(gt_text, pairs, chunk_embs, gt_emb)

        # ndcg
        nd = ndcg_at_k(retrieved, gt_chunk, top_k)

        # sim stats
        topk_embs = chunk_embs[retrieved]
        sim_topk = np.dot(topk_embs, q_emb)
        rows.append({
            "id": row["id"],
            "question": row["question"],
            "gt_chunk_id": gt_chunk+1,
            "retrieved_ids": [int(r)+1 for r in retrieved],
            "hit_at_k": int(gt_chunk in retrieved),
            "cluster_hit@k": cl_hit,
            "text_hit@k": text_hit,
            "exact_score": exact_sc,
            "overlap_score": overlap_sc,
            "semantic_score": sem_sc,
            "rank": rank,
            "rr": rr,
            "ndcg@k": nd,
            "topk_sim_max": float(np.max(sim_topk)),
            "topk_sim_avg": float(np.mean(sim_topk)),
        })

    df_res = pd.DataFrame(rows)
    df_res["alignment_status"] = df_res.apply(detect_misalignment,axis=1)

    misaligned = df_res["alignment_status"].isin(["SHIFTED","SHIFTED_NEAR"]).mean()

    summary = {
        "MRR": df_res["rr"].mean(),
        "Recall@K": df_res["hit_at_k"].mean(),
        "ClusterRecall@K": df_res["cluster_hit@k"].mean(),
        "RankAvg": df_res["rank"].mean(),
        "nDCG@K": df_res["ndcg@k"].mean(),
        "TopKSimMaxAvg": df_res["topk_sim_max"].mean(),
        "TopKSimAvg": df_res["topk_sim_avg"].mean(),
        "TextHit@K": df_res["text_hit@k"].mean(),
        "MisalignmentRate": misaligned,
        "ExactScoreAvg": df_res["exact_score"].mean(),
        "OverlapScoreAvg": df_res["overlap_score"].mean(),
        "SemanticScoreAvg": df_res["semantic_score"].mean(),
    }

    bloom_summary = pd.DataFrame()
    if "BloomLevel" in df.columns:
        df_res["BloomLevel"] = df["BloomLevel"]
        bloom_summary = (
            df_res.groupby("BloomLevel")
            .agg({
                "rr":"mean","hit_at_k":"mean","text_hit@k":"mean","cluster_hit@k":"mean",
                "rank":"mean","ndcg@k":"mean","topk_sim_max":"mean","topk_sim_avg":"mean",
                "exact_score":"mean","overlap_score":"mean","semantic_score":"mean"
            })
            .rename(columns={
                "rr":"MRR","hit_at_k":"Recall@K","text_hit@k":"TextHit@K","cluster_hit@k":"ClusterRecall@K",
                "rank":"RankAvg","ndcg@k":"nDCG@K","topk_sim_max":"TopKSimMaxAvg","topk_sim_avg":"TopKSimAvg",
                "exact_score":"ExactScoreAvg","overlap_score":"OverlapScoreAvg","semantic_score":"SemanticScoreAvg"})
            .reset_index()
        )

    return df_res, chunks, summary, bloom_summary, full_sim_matrix, index, chunk_embs

# ============================================================
# Excel
# ============================================================
def save_excel(exp_id, out, df_res, chunks, summary, bloom, heatmap, scatter):
    p = f"{out}/{exp_id}.xlsx"
    df_sum = pd.DataFrame({"metric":summary.keys(), "value":summary.values()})
    df_chunks = pd.DataFrame({
        "chunk_id": range(1,len(chunks)+1),
        "chunk_text": chunks,
        "tokens": [len(c.split()) for c in chunks],
        "chars": [len(c) for c in chunks]
    })
    with pd.ExcelWriter(p, engine="openpyxl") as w:
        df_res.to_excel(w, index=False, sheet_name="retrieval")
        format_decimal(w.sheets["retrieval"])
        df_sum.to_excel(w, index=False, sheet_name="summary")
        format_decimal(w.sheets["summary"])
        if bloom is not None and len(bloom)>0:
            bloom.to_excel(w, index=False, sheet_name="bloom")
            format_decimal(w.sheets["bloom"])
        df_chunks.to_excel(w, index=False, sheet_name="chunks")
        format_decimal(w.sheets["chunks"])
        wb = w.book
        ws_h = wb.create_sheet("heatmap"); img_h = XLImage(heatmap); img_h.anchor="A1"; ws_h.add_image(img_h)
        ws_s = wb.create_sheet("scatter"); img_s = XLImage(scatter); img_s.anchor="A1"; ws_s.add_image(img_s)

# ============================================================
# MAIN
# ============================================================
def main():
    print("\n==========================================")
    print("🚀 STARTING ADVANCED RETRIEVAL PIPELINE")
    print("==========================================")

    cfg = load_yaml("configs/retrieval_adv.yaml")
    master = cfg["master"]
    exps = cfg["experiments"]

    print(f"📌 Loaded config  : configs/retrieval_adv.yaml")
    print(f"📌 PDF directory  : {master['pdf_dir']}")
    print(f"📌 Dataset        : {master['dataset']}")
    print("------------------------------------------")

    df = pd.read_excel(master["dataset"])
    print(f"📄 Dataset loaded : {len(df)} rows")
    df = pd.read_excel(master["dataset"])
    print(f"📄 Dataset loaded : {len(df)} rows")

    # ---------------------------------------------
    # Preprocessing untuk Bloom dataset
    # flatten dict into clean text
    # ---------------------------------------------
    import ast

    def flatten_context(x):
        try:
            d = ast.literal_eval(x)
            if isinstance(d, dict):
                return " ".join(v.strip() for v in d.values())
            return x
        except:
            return x

    df["context"] = df["context"].astype(str).apply(flatten_context)
    #----------------------------------------------
    global_summary = {}
    all_bloom = []

    for exp in exps:
        if not exp.get("active", False):
            print(f"⏭ Skipping experiment '{exp['id']}' (inactive)")
            continue

        eid = exp["id"]
        print("\n------------------------------------------")
        print(f"▶ Running experiment: {eid}")
        print("------------------------------------------")

        out = auto_out(eid)
        print(f"📂 Output directory: {out}")

        # -------------------------------
        # Run retrieval + evaluation
        # -------------------------------
        print("🔍 Running retrieval + evaluator ...")
        df_res, chunks, summary, bloom, sim_mat, index, chunk_embs = run_retrieval_adv(exp, master, df)

        print(f"✔ Retrieval done | Retrieved chunks: {len(chunks)}")

        # -------------------------------
        # Visualizations
        # -------------------------------
        print("📊 Generating heatmap ...")
        heatmap = plot_heatmap(sim_mat, out, eid)

        print("📈 Generating scatter plot ...")
        scatter = plot_scatter_rank_similarity(df_res, out)

        # -------------------------------
        # Save FAISS metadata
        # -------------------------------
        print("💾 Saving FAISS index + embeddings ...")
        save_faiss(out, index, chunk_embs, chunks)

        # -------------------------------
        # Save Excel
        # -------------------------------
        print("📘 Writing Excel report ...")
        save_excel(eid, out, df_res, chunks, summary, bloom, heatmap, scatter)

        # -------------------------------
        # Global summary collection
        # -------------------------------
        global_summary[eid] = summary
        if len(bloom) > 0:
            bloom["experiment_id"] = eid
            all_bloom.append(bloom)

        print(f"✔ Experiment '{eid}' completed.")

    # ============================================================
    # Global outputs
    # ============================================================
    summary_dir = f"outputsRetrievalADV/{stamp}"
    ensure_dir(summary_dir)

    print("\n==========================================")
    print("📦 Saving global summaries ...")
    print("==========================================")

    df_global = (
        pd.DataFrame.from_dict(global_summary, orient="index")
        .reset_index()
        .rename(columns={"index": "experiment_id"})
    )

    df_global.to_excel(
        f"{summary_dir}/summary_all_{stamp}.xlsx",
        index=False
    )
    print(f"✔ Global summary saved: {summary_dir}/summary_all_{stamp}.xlsx")

    if all_bloom:
        df_bloom_global = pd.concat(all_bloom, ignore_index=True)
        df_bloom_global.to_excel(
            f"{summary_dir}/summary_bloom_all_{stamp}.xlsx",
            index=False
        )
        print(f"✔ Bloom analysis summary saved: {summary_dir}/summary_bloom_all_{stamp}.xlsx")

    print("\n==========================================")
    print(f"🎉 ALL EXPERIMENTS COMPLETED")
    print(f"📁 Output stored in: {summary_dir}")
    print("==========================================\n")

if __name__=="__main__":
    main()
