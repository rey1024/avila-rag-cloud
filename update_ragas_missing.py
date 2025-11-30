import pandas as pd
import numpy as np
import logging
from utils_generation.eval_ragas import run_ragas_evaluation
from openpyxl import load_workbook
import os
from datetime import datetime

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)
# ===============================================================
# 0. LOAD ENV
# ===============================================================
from dotenv import load_dotenv
load_dotenv()
print("✔ Env loaded from .env")

# ===================================================================
# Helper untuk menyalin sheet lama ke writer baru
# ===================================================================
def copy_existing_sheet(writer, wb_old, sheet_name):
    if sheet_name not in wb_old.sheetnames:
        return

    ws_old = wb_old[sheet_name]
    ws_new = writer.book.create_sheet(sheet_name)

    for row in ws_old.iter_rows():
        for cell in row:
            ws_new[cell.coordinate].value = cell.value


# ===================================================================
# Fungsi utama: memperbarui skor RAGAS lalu simpan ke file baru
# ===================================================================
def update_ragas_missing(
    excel_path,
    openai_key,
):
    """
    - Membaca file excel
    - Cek baris yang metrik RAGAS-nya kosong
    - Hitung ulang hanya baris kosong
    - Buat FILE BARU: <nama>_v-update_YYYYMMDD_HHMMSS.xlsx
    - Isi semua sheet, update results, summary dihitung ulang
    """

    logger.info("======================================================")
    logger.info(f"📂 Loading Excel: {excel_path}")
    logger.info("======================================================")

    # -----------------------------------------------
    # Load workbook lama (semua sheet)
    # -----------------------------------------------
    wb_old = load_workbook(excel_path)
    df_results = pd.read_excel(excel_path, sheet_name="results")

    metric_cols = [
        "context_precision",
        "context_recall",
        "answer_relevancy",
        "faithfulness",
        "answer_correctness"
    ]

    missing_mask = df_results[metric_cols].isnull().any(axis=1)
    missing_indices = df_results[missing_mask].index.tolist()

    logger.info(f"🔎 Ada {len(missing_indices)} baris yang perlu dihitung ulang.")

    # ------------------------------------------------
    # Tidak ada baris kosong → tetap buat file update
    # ------------------------------------------------
    questions, answers, contexts, gts, row_map = [], [], [], [], []

    for idx in missing_indices:
        row = df_results.iloc[idx]

        questions.append(row["question"])
        answers.append(row["llm_answer"])
        contexts.append(row["contexts_retrieval"].split("\n"))
        gts.append(row["answer"])
        row_map.append(idx)

        logger.info(f"• Re-eval Row {idx}: {row['question'][:60]}...")

    # ------------------------------------------------
    # Jika ada yang harus dievaluasi ulang → jalankan RAGAS
    # ------------------------------------------------
    if len(row_map) > 0:
        ragas_items, ragas_agg_partial = run_ragas_evaluation(
            questions,
            answers,
            contexts,
            gts,
            openai_key
        )

        # update baris
        for local_i, excel_row in enumerate(row_map):
            item = ragas_items[local_i]

            df_results.loc[excel_row, "context_precision"] = item["context_precision"]
            df_results.loc[excel_row, "context_recall"] = item["context_recall"]
            df_results.loc[excel_row, "answer_relevancy"] = item["answer_relevancy"]
            df_results.loc[excel_row, "faithfulness"] = item["faithfulness"]
            df_results.loc[excel_row, "answer_correctness"] = item["answer_correctness"]

            logger.info(
                f"✔ Updated row {excel_row}: "
                f"CP={item['context_precision']}, "
                f"CR={item['context_recall']}, "
                f"AR={item['answer_relevancy']}, "
                f"F={item['faithfulness']}, "
                f"AC={item['answer_correctness']}"
            )
    else:
        logger.info("✔ Tidak ada skor kosong. Semuanya lengkap.")


    # ------------------------------------------------
    # Hitung ulang summary global
    # ------------------------------------------------
    df_summary = pd.DataFrame({
        "metric": metric_cols,
        "score": [
            df_results["context_precision"].mean(),
            df_results["context_recall"].mean(),
            df_results["answer_relevancy"].mean(),
            df_results["faithfulness"].mean(),
            df_results["answer_correctness"].mean()
        ]
    })

    # ------------------------------------------------
    # Buat nama file baru
    # ------------------------------------------------
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(excel_path)

    new_excel_path = f"{base}_v-update_{timestamp}{ext}"

    logger.info("======================================================")
    logger.info(f"💾 Menyimpan file baru → {new_excel_path}")
    logger.info("======================================================")

    # ------------------------------------------------
    # Tulis file baru
    # ------------------------------------------------
    with pd.ExcelWriter(new_excel_path, engine="openpyxl") as writer:

        # 1. tulis summary baru
        df_summary.to_excel(writer, index=False, sheet_name="summary")

        # 2. tulis results updated
        df_results.to_excel(writer, index=False, sheet_name="results")

        # 3. salin sheet lain dari file lama jika ada
        for sheet_name in wb_old.sheetnames:
            if sheet_name in ["summary", "results"]:
                continue
            copy_existing_sheet(writer, wb_old, sheet_name)

    logger.info("✔ Semua data berhasil disimpan.")
    return new_excel_path


# ============================================================
# Contoh pemanggilan
# ============================================================
if __name__ == "__main__":
    OPENAI_KEY = os.getenv("OPENAI_API_KEY")
    update_ragas_missing("results/generation/Gen20251129_194615/GEN_GPT4oMini_20251129_194615/GEN_GPT4oMini_20251129_194615.xlsx", OPENAI_KEY)
