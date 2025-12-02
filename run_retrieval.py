import os
import yaml
import faiss
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import json

from pathlib import Path
from datetime import datetime
from sentence_transformers import SentenceTransformer
from nltk.tokenize import sent_tokenize
import openpyxl
import nltk
nltk.download('punkt')
nltk.download('punkt_tab')
stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
from dotenv import load_dotenv
load_dotenv()
import os
import sys
from sentence_transformers import CrossEncoder

# memastikan folder project menjadi root import
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)
from utils_common.utils import *
import logging
OUTPUT_DIR=f"results/retrieval/ret_{stamp}"
setup_logging(OUTPUT_DIR)
logger = logging.getLogger(__name__)



# ================================================
# BLOOM LEVEL CLASSIFIER  (GPT 4o mini)
# ================================================
from openai import OpenAI
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
def short_text(t, max_len=200):
    t = str(t)
    return t if len(t) <= max_len else t[:max_len] + " ... [truncated]"

def classify_bloom_gpt(paragraph):
    logger.info(f"Classifying paragraph with GPT (length: {len(paragraph)} chars)")
    prompt = f"""
    Klasifikasikan paragraf berikut ke salah satu level Taksonomi Bloom:
    Remember, Understand, Apply, Analyze, Evaluate, Create.
    Jawab hanya satu kata.

    Paragraf:
    {paragraph}
    """

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0
        )
        label = response.choices[0].message['content'].strip().lower()
        logger.info(f"GPT classification result: {label}")
    except Exception as e:
        logger.warning(f"GPT classification failed: {e}, defaulting to remember")
        label = "remember"

    mapping = {
        "remember": "Remember",
        "understand": "Understand",
        "apply": "Apply",
        "analyze": "Analyze",
        "evaluate": "Evaluate",
        "create": "Create"
    }

    final_label = mapping.get(label, "Remember")
    logger.info(f"Mapped label: {final_label}")
    return final_label


# ================================================
# BLOOM-AWARE CHUNKING
# ================================================
def bloom_aware_chunk(text, min_tokens=40, max_tokens=280):
    logger.info(f"Bloom aware chunking with min_tokens: {min_tokens}, max_tokens: {max_tokens}")
    paragraphs = [p.strip() for p in text.split("\n") if len(p.strip()) > 0]
    total = len(paragraphs)

    logger.info(f"[BloomChunk] Total paragraphs: {total}")

    bloom_buckets = {
        "Remember": [], "Understand": [], "Apply": [],
        "Analyze": [], "Evaluate": [], "Create": []
    }

    for i, p in enumerate(paragraphs, start=1):
        logger.info(f"[BloomChunk] {i}/{total} → classifying...")

        try:
            label = classify_bloom_gpt(p)
            logger.info(f"[BloomChunk] Paragraph {i} classified as: {label}")
        except Exception as e:
            logger.warning(f"[BloomChunk] Classification failed for paragraph {i}: {e}, defaulting to Remember")
            label = "Remember"

        token_len = len(p.split())

        if token_len > max_tokens:
            logger.info(f"[BloomChunk] Paragraph {i} too long ({token_len} tokens), splitting")
            mid = len(p)//2
            bloom_buckets[label].append(p[:mid])
            bloom_buckets[label].append(p[mid:])
        else:
            bloom_buckets[label].append(p)

    logger.info("[BloomChunk] Classification complete. Assembling Bloom chunks...")

    # gabungkan chunk berdasarkan level bloom
    final_chunks = []
    final_metadata = []

    chunk_id = 1
    for bloom_level, paras in bloom_buckets.items():
        if len(paras) == 0:
            continue

        chunk_text = "\n".join(paras)
        final_chunks.append(chunk_text)

        meta = {
            "chunk_id": chunk_id,
            "bloom_level": bloom_level,
            "paragraph_count": len(paras),
            "tokens": len(chunk_text.split()),
            "text_preview": chunk_text[:200]
        }
        final_metadata.append(meta)
        logger.info(f"[BloomChunk] Created chunk {chunk_id} for {bloom_level}: {len(paras)} paras, {meta['tokens']} tokens")
        chunk_id += 1

    logger.info(f"[BloomChunk] Total final chunks: {len(final_chunks)}")
    return final_chunks, final_metadata

# ================================================
# LOAD PDF/DOCX
# ================================================
def load_pdf_docx(folder):
    logger.info(f"Loading PDF/DOCX from folder: {folder}")
    import pypdf
    from docx import Document
    path = Path(folder)

    pdfs = list(path.glob("*.pdf"))
    docs = list(path.glob("*.docx"))
    logger.info(f"Found {len(pdfs)} PDF files and {len(docs)} DOCX files")

    texts = []

    for p in pdfs:
        logger.info(f"Processing PDF: {p}")
        reader = pypdf.PdfReader(str(p))
        t = " ".join(page.extract_text() or "" for page in reader.pages)
        texts.append(t)
        logger.info(f"Extracted {len(t)} chars from {p}")

    for d in docs:
        logger.info(f"Processing DOCX: {d}")
        doc = Document(str(d))
        paragraphs = [p.text for p in doc.paragraphs if p.text]
        t = "\n".join(paragraphs)
        texts.append(t)
        logger.info(f"Extracted {len(paragraphs)} paragraphs from {d}")

    if not texts:
        logger.error("No PDF/DOCX files found")
        raise ValueError("Tidak ada PDF/DOCX ditemukan")

    combined_text = "\n".join(texts)
    logger.info(f"Combined text length: {len(combined_text)} chars")
    return combined_text

def format_decimal(ws):
    for row in ws.iter_rows():
        for cell in row:
            try:
                # Coba konversi ke float
                val = float(cell.value)
                cell.number_format = "0.00"
            except:
                pass



# ================================================
# YAML UTILS
# ================================================
def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def auto_out(exp_id):
    
    path = f"{OUTPUT_DIR}/{exp_id}_{stamp}"
    ensure_dir(path)
    return path


# ================================================
# CHUNKING
# ================================================
def fixed_chunk(text, size, overlap):
    tokens = text.split()
    step = size - overlap
    chunks = []
    for i in range(0, len(tokens), step):
        seg = tokens[i:i + size]
        if seg:
            chunks.append(" ".join(seg))
    return chunks


def semantic_chunk_langchain(text, embedder, 
                             min_chunk_size=150, 
                             max_chunk_size=450,
                             similarity_threshold=0.75):

    # 1. Tokenisasi kalimat
    sents = sent_tokenize(text)
    if len(sents) == 0:
        return []

    # 2. Embedding semua kalimat
    sent_embs = embedder.encode(sents, normalize_embeddings=True, batch_size=64)
    # Hitung similarity antar kalimat berdekata
    pair_sims = []
    for i in range(1, len(sent_embs)):
        pair_sims.append(np.dot(sent_embs[i], sent_embs[i - 1]))

    # threshold dinamis => persentil 30%
    dynamic_thr = np.percentile(pair_sims, 30)

    # pakai threshold kecil antara preset dan dinamis
    threshold_final = min(similarity_threshold, dynamic_thr)

    chunks = []
    current_chunk = []
    current_len = 0

    for i in range(len(sents)):
        sent = sents[i]
        sent_len = len(sent.split())

        # Jika chunk masih kosong, mulai chunk baru
        if not current_chunk:
            current_chunk.append(sent)
            current_len += sent_len
            continue

        # 3. Hitung similarity ke kalimat sebelumnya
        # sim = np.dot(sent_embs[i], sent_embs[i - 1])

        # # Jika similarity turun → boundary potensial
        # boundary = sim < similarity_threshold
        # similarity ke kalimat sebelumnya
        sim_prev = np.dot(sent_embs[i], sent_embs[i - 1])

        # similarity ke seluruh chunk yang sudah terbentuk
        current_chunk_emb = np.mean(sent_embs[max(0, i-len(current_chunk)):i], axis=0)
        sim_chunk = np.dot(sent_embs[i], current_chunk_emb)

        # dua threshold: untuk kalimat sebelumnya dan chunk akumulatif
        boundary = (sim_prev < threshold_final) and (sim_chunk < (threshold_final - 0.10))


        # Jika ukuran chunk sudah besar → wajib boundary
        too_big = current_len + sent_len > max_chunk_size

        # Jika boundary atau chunk terlalu panjang → simpan chunk
        if boundary or too_big:
            # Jika chunk terlalu kecil → jangan dipotong
            if current_len < min_chunk_size and not too_big:
                current_chunk.append(sent)
                current_len += sent_len
                continue

            # Simpan chunk
            chunks.append(" ".join(current_chunk))
            current_chunk = [sent]
            current_len = sent_len
        else:
            # Tambahkan kalimat
            current_chunk.append(sent)
            current_len += sent_len

    # Sisa chunk terakhir
    if current_chunk:
        chunks.append(" ".join(current_chunk))

    return chunks


def chunk_text(text, cfg, embedder=None):
    logger.info(f"Chunking text with method: {cfg['type']}, config: {cfg}")
    # FIXED
    if cfg["type"] == "fixed":
        logger.info(f"Fixed chunking with size: {cfg['size']}, overlap: {cfg['overlap']}")
        chunks = fixed_chunk(text, cfg["size"], cfg["overlap"])
        logger.info(f"Fixed chunks: {len(chunks)}")
        return chunks, None

    # SENTENCE
    if cfg["type"] == "sentence":
        logger.info(f"Sentence chunking with window: {cfg['window']}, overlap: {cfg['overlap']}")
        chunks = fixed_chunk(text, cfg["window"], cfg["overlap"])
        logger.info(f"Sentence chunks: {len(chunks)}")
        return chunks, None

    # SEMANTIC
    if cfg["type"] == "semantic":
        logger.info(f"Semantic chunking with min: {cfg.get('min_chunk_size', 150)}, max: {cfg.get('max_chunk_size', 450)}, threshold: {cfg.get('threshold', 0.75)}")
        chunks = semantic_chunk_langchain(
            text,
            embedder,
            min_chunk_size=cfg.get("min_chunk_size", 150),
            max_chunk_size=cfg.get("max_chunk_size", 450),
            similarity_threshold=cfg.get("threshold", 0.75),
        )
        logger.info(f"Semantic chunks: {len(chunks)}")
        return chunks, None

    # BLOOM LLM
    if cfg["type"] == "bloom_llm":
        logger.info(f"Bloom LLM chunking with min_tokens: {cfg.get('min_chunk_size', 40)}, max_tokens: {cfg.get('max_chunk_size', 280)}")
        chunks, metadata = bloom_aware_chunk(
            text,
            min_tokens=cfg.get("min_chunk_size", 40),
            max_tokens=cfg.get("max_chunk_size", 280)
        )
        logger.info(f"Bloom chunks: {len(chunks)}")
        return chunks, metadata

    logger.error(f"Unknown chunking method: {cfg['type']}")
    raise ValueError("Unknown chunking method")

# ======================================================
# BOOSTED GROUNDTRUTH MATCHING UTILITIES
# ======================================================

def token_overlap(a, b):
    sa = set(a.split())
    sb = set(b.split())
    if len(sa) == 0 or len(sb) == 0:
        return 0
    return len(sa & sb) / len(sa | sb)


def map_groundtruth_chunk(context_text, chunks, chunk_embs, gt_emb):
    # -------- METHOD 1: Exact containment --------
    contain_candidates = []
    for i, ch in enumerate(chunks):
        if context_text.strip() in ch:
            contain_candidates.append(i)

    if contain_candidates:
        return contain_candidates[0]

    # -------- METHOD 2: Token Overlap --------
    overlaps = [token_overlap(context_text, c) for c in chunks]
    overlap_best = int(np.argmax(overlaps))
    overlap_score = overlaps[overlap_best]

    # -------- METHOD 3: Semantic Similarity --------
    sims = np.dot(chunk_embs, gt_emb)
    sem_best = int(np.argmax(sims))
    sem_score = sims[sem_best]

    # -------- RULE-BASED DECISION --------
    if overlap_score > 0.15:
        return overlap_best

    return sem_best

def detect_misalignment(row):
    hit = row["hit_at_k"]
    cluster = row["cluster_hit@k"]
    text_hit = row["text_hit@k"]

    # Tidak ditemukan sama sekali
    if text_hit == 0:
        return "NOT_FOUND"

    # ID match sempurna
    if hit == 1:
        return "ALIGNED"

    # Cluster dekat (±1)
    if cluster == 1:
        return "SHIFTED_NEAR"

    # Teks match tetapi ID tidak ditemukan sama sekali → shifted jauh
    if hit == 0 and text_hit == 1:
        return "SHIFTED"

    return "UNKNOWN"

def cluster_neighbors(idx, total, radius=1):
    neighbors = set()
    for r in range(-radius, radius + 1):
        j = idx + r
        if 0 <= j < total:
            neighbors.add(j)
    return neighbors


def ndcg_at_k(retrieved_ids, gt_id, k):
    gains = [1 if rid == gt_id else 0 for rid in retrieved_ids[:k]]
    dcg = 0

    if gains:
        dcg += gains[0]
        for i, g in enumerate(gains[1:], start=2):
            dcg += g / np.log2(i)

    ideal = 1  # because only one relevant item exists
    return dcg / ideal

# ======================================================
# TEXT-BASED HIT MATCHING (exact, overlap, semantic)
# ======================================================
def text_hit_match(
    gt_text, 
    retrieved_chunks, 
    chunk_embs, 
    gt_emb,
    overlap_threshold=0.35,      # lebih ketat sedikit
    semantic_threshold=0.90,     # balanced precision
    exact_window=200             # batasi exact match
):
    """
    Versi balanced:
    - Tidak boleh semantic-only hit
    - Harus ada exact match kecil ATAU overlap tinggi
    - Ditambah semantic similarity yang kuat
    """

    best_exact = 0
    best_overlap = 0
    best_sim = 0

    for rid, chunk_text in retrieved_chunks:

        # EXACT match ketat (hanya 200 karakter pertama)
        if (
            gt_text[:exact_window].strip() in chunk_text
            or chunk_text[:exact_window].strip() in gt_text
        ):
            best_exact = 1

        # OVERLAP
        ov = token_overlap(gt_text, chunk_text)
        best_overlap = max(best_overlap, ov)

        # SEMANTIC
        sim = float(np.dot(chunk_embs[rid], gt_emb))
        best_sim = max(best_sim, sim)

    # 1) Semantic harus kuat
    semantic_ok = best_sim >= semantic_threshold

    # 2) Exact match kecil atau overlap lumayan
    text_match_ok = (best_exact == 1) or (best_overlap >= overlap_threshold)

    # FINAL HIT rule (balanced):
    # - semantic kuat
    # - ditambah exact OR overlap yang cukup
    hit = 1 if (semantic_ok and text_match_ok) else 0

    return hit, best_exact, best_overlap, best_sim


# ================================================
# RETRIEVAL PIPELINE
# ================================================
# ======================================================
# RETRIEVAL PIPELINE (FINAL VERSION)
# ======================================================

def run_retrieval(exp, master, df, out_dir):
    logger.info(f"\n=== Running experiment {exp['id']} ===")
    logger.info(f"Experiment config: {exp}")
    logger.info(f"Master config: {master}")
    logger.info(f"Dataset shape: {df.shape}")
    logger.info(f"Output dir: {out_dir}")

    reranker = CrossEncoder("BAAI/bge-reranker-v2-m3", device="cuda")
    logger.info("Loaded CrossEncoder reranker")

    import torch
    torch.cuda.empty_cache()
    torch.cuda.ipc_collect()
    logger.info("[GPU] Cache cleared")

    embedder = SentenceTransformer(
        exp["embedding"]["model"],
        device="cuda",
        trust_remote_code=True
    )
    logger.info(f"Loaded embedder: {exp['embedding']['model']}")

    corpus_text = load_pdf_docx(master["pdf_dir"])
    logger.info(f"Loaded corpus text from {master['pdf_dir']}, length: {len(corpus_text)} chars")

    chunks, bloom_metadata = chunk_text(corpus_text, exp["chunking"], embedder)
    logger.info(f"Chunked text into {len(chunks)} chunks")

    # save bloom metadata if exists
    if bloom_metadata is not None:
        bloom_meta_path = f"{out_dir}/chunk_bloom_metadata.jsonl"
        with open(bloom_meta_path, "w", encoding="utf-8") as f:
            for m in bloom_metadata:
                f.write(json.dumps(m, ensure_ascii=False) + "\n")
        logger.info(f"Saved bloom metadata to {bloom_meta_path}")

    logger.info("Starting chunk embedding...")
    chunk_embs = embedder.encode(chunks, normalize_embeddings=True, batch_size=64)
    logger.info(f"Chunk embeddings shape: {chunk_embs.shape}")

    logger.info("Starting question embedding...")
    q_embs = embedder.encode(df["question"].tolist(), normalize_embeddings=True, batch_size=64)
    logger.info(f"Question embeddings shape: {q_embs.shape}")

    logger.info("Starting ground truth embedding...")
    gt_embs = embedder.encode(df["context"].tolist(), normalize_embeddings=True, batch_size=64)
    logger.info(f"Ground truth embeddings shape: {gt_embs.shape}")

    dim = chunk_embs.shape[1]
    index = faiss.IndexFlatIP(dim)
    index.add(chunk_embs)

    # res = faiss.StandardGpuResources()
    # index = faiss.index_cpu_to_gpu(res, 0, faiss.IndexFlatL2(dim))
    # index.add(chunk_embs)
    logger.info(f"Built FAISS index with {index.ntotal} vectors, dim: {dim}")
    logger.info(f"Built FAISS index with {index.ntotal} vectors, dim: {dim}")

    top_k = master.get("top_k",10)
    logger.info(f"Top K: {top_k}")
    rows = []

    logger.info("Computing full similarity matrix...")
    #full_sim_matrix = np.dot(chunk_embs, q_embs.T)
    
    chunk_t = torch.tensor(chunk_embs, device="cuda")
    q_t = torch.tensor(q_embs, device="cuda")
    full_sim_matrix = torch.matmul(chunk_t, q_t.T).cpu().numpy()
    logger.info(f"Similarity matrix shape: {full_sim_matrix.shape}")

    logger.info("Starting retrieval loop...")
    
    
    #res = faiss.StandardGpuResources()
    #gpu_index = faiss.index_cpu_to_gpu(res, 0, faiss.IndexFlatL2(dim))
    #gpu_index.add(chunk_embs)    
    for i, row in df.iterrows():
        logger.info(f"Processing row {i+1}/{len(df)}: {row['id']}")
        q_emb = q_embs[i]
        gt_emb = gt_embs[i]
        gt_text = row["context"]
        # === LOG: Question & GT Context ===
        logger.info("\n------------------------------------------")
        logger.info(f"[Row {i}] QUESTION:")
        logger.info(row["question"])
        logger.info("\n[Row {i}] GROUNDTRUTH CONTEXT:")
        logger.info(gt_text)
        logger.info("------------------------------------------\n\n")

        # ------------------------------------------------
        # ROBUST GROUNDTRUTH CHUNK MATCHING
        # ------------------------------------------------
        gt_chunk = map_groundtruth_chunk(gt_text, chunks, chunk_embs, gt_emb)
        logger.info(f"Row {i}: GT chunk: {gt_chunk}")
        logger.info(f"[Row {i}] GT CHUNK ID: {gt_chunk}")
        logger.info(f"[Row {i}] GT CHUNK TEXT:\n{short_text(chunks[gt_chunk])}")
        logger.info(f"GT_TEXT: {short_text(gt_text)}")


        # ------------------------------------------------
        # RETRIEVAL
        # ------------------------------------------------
        logger.info(f"Row {i}: Performing FAISS search...")
    


        #scores, idx = gpu_index.search(q_emb.reshape(1, -1), top_k)
        scores, idx = index.search(q_emb.reshape(1, -1), top_k)



        #scores, idx = index.search(q_emb.reshape(1, -1), top_k)
        retrieved = idx[0]
        logger.info(f"Row {i}: Initial retrieved: {retrieved}")

        # Rerank top_k with cross-encoder
        logger.info(f"Row {i}: Reranking with cross-encoder...")
        rerank_inputs = [[row["question"], chunks[rid]] for rid in retrieved]
        #rerank_scores = reranker.predict(rerank_inputs)
        rerank_scores = reranker.predict(rerank_inputs, batch_size=32)

        logger.info(f"Row {i}: Rerank scores: {rerank_scores}")

        # urutkan berdasarkan score cross encoder
        reranked = [retrieved[i] for i in np.argsort(rerank_scores)[::-1]]
        retrieved = np.array(reranked)
        logger.info(f"Row {i}: Final retrieved after reranking: {retrieved}")
        logger.info(f"[Row {i}] === RETRIEVED CHUNKS (Top-{top_k}) ===")

        for rank_num, chunk_id in enumerate(retrieved, start=1):
            logger.info(f"\n[Rank {rank_num}] Chunk ID: {chunk_id}")
            logger.info(f"[Rank {rank_num}] TEXT:\n{short_text(chunks[chunk_id])}")


            # ------------------------------------------------
            # TEXT HIT K (exact + overlap + semantic similarities)
            # ------------------------------------------------
            # Pair: (chunk_id, chunk_text) untuk setiap retrieved
            retrieved_pairs = [(rid, chunks[rid]) for rid in retrieved]

            text_hit, exact_score, overlap_score, semantic_score = text_hit_match(
                gt_text=gt_text,
                retrieved_chunks=retrieved_pairs,
                chunk_embs=chunk_embs,
                gt_emb=gt_emb,
                overlap_threshold=0.25,
                semantic_threshold=0.80
            )
            logger.info(f"Row {i}: Text hit: {text_hit}, exact: {exact_score}, overlap: {overlap_score}, semantic: {semantic_score}")
            logger.info(f"[Row {i}] FINAL TEXT_HIT@K = {text_hit}")
            logger.info("------------------------------------------\n\n")


            # ------------------------------------------------
            # RANKING
            # ------------------------------------------------
            # sims_q = np.dot(chunk_embs, q_emb)
            # rank_order = np.argsort(sims_q)[::-1]
            # rank = int(np.where(rank_order == gt_chunk)[0][0]) + 1
            # rr = 1 / rank
            # logger.info(f"Row {i}: Rank: {rank}, RR: {rr}")
            if gt_chunk in retrieved:
                rank = list(retrieved).index(gt_chunk) + 1
            else:
                rank = top_k + 1
            rr = 1.0 / rank

            
            # ------------------------------------------------
            # CLUSTER-BASED HIT (CHUNK ±1)
            # ------------------------------------------------
            cluster = cluster_neighbors(gt_chunk, len(chunks), radius=1)
            cluster_hit = int(any([rid in cluster for rid in retrieved]))
            logger.info(f"Row {i}: Cluster hit: {cluster_hit}, cluster: {cluster}")

            # ------------------------------------------------
            # nDCG@K
            # ------------------------------------------------
            ndcg = ndcg_at_k(retrieved, gt_chunk, top_k)
            logger.info(f"Row {i}: nDCG@K: {ndcg}")

            # ------------------------------------------------
            # SIMILARITY STATS
            # ------------------------------------------------
            topk_embs = chunk_embs[retrieved]
            topk_sims = np.dot(topk_embs, q_emb)
            logger.info(f"Row {i}: TopK sim max: {np.max(topk_sims)}, avg: {np.mean(topk_sims)}")

            rows.append({
                "id": row["id"],
                "question": row["question"],
                "groundtruth_context": gt_text,
                "retrieved_chunks_text": " ||| ".join([chunks[rid] for rid in retrieved]),
                "gt_chunk_id": gt_chunk + 1,
                "retrieved_ids": [r + 1 for r in retrieved],
                "hit_at_k": int(gt_chunk in retrieved),
                "cluster_hit@k": cluster_hit,
                "text_hit@k": text_hit,

                # New detailed scores
                "exact_score": exact_score,
                "overlap_score": overlap_score,
                "semantic_score": semantic_score,

                "rank": rank,
                "rr": rr,
                "ndcg@k": ndcg,
                "topk_sim_max": float(np.max(topk_sims)),
                "topk_sim_avg": float(np.mean(topk_sims)),
            })


    logger.info("Creating results DataFrame...")
    df_res = pd.DataFrame(rows)
    logger.info(f"Results DataFrame shape: {df_res.shape}")

    logger.info("Detecting misalignment...")
    df_res["alignment_status"] = df_res.apply(detect_misalignment, axis=1)
    misaligned_mask = df_res["alignment_status"].isin(["SHIFTED", "SHIFTED_NEAR"])
    misalignment_rate = misaligned_mask.mean()
    logger.info(f"Misalignment rate: {misalignment_rate}")

    logger.info("Computing summary metrics...")
    summary = {
        "MRR": df_res["rr"].mean(),
        "Recall@K": df_res["hit_at_k"].mean(),
        "ClusterRecall@K": df_res["cluster_hit@k"].mean(),
        "RankAvg": df_res["rank"].mean(),
        "nDCG@K": df_res["ndcg@k"].mean(),
        "TopKSimMaxAvg": df_res["topk_sim_max"].mean(),
        "TopKSimAvg": df_res["topk_sim_avg"].mean(),
        "TextHit@K": df_res["text_hit@k"].mean(),
        "MisalignmentRate": misalignment_rate,
        "TrueRetrievalSuccess": df_res["text_hit@k"].mean(),
        "ExactScoreAvg": df_res["exact_score"].mean(),
        "OverlapScoreAvg": df_res["overlap_score"].mean(),
        "SemanticScoreAvg": df_res["semantic_score"].mean(),
        }
    logger.info(f"Summary: {summary}")



    # ============================================================
    # BLOOM LEVEL ANALYSIS
    # ============================================================
    if "BloomLevel" in df.columns:
        logger.info("BloomLevel column found, performing bloom analysis")
        df_res["BloomLevel"] = df["BloomLevel"]

        bloom_summary = (
            df_res.groupby("BloomLevel")
            .agg({
                "rr": "mean",
                "hit_at_k": "mean",
                "text_hit@k": "mean",
                "cluster_hit@k": "mean",
                "rank": "mean",
                "ndcg@k": "mean",
                "topk_sim_max": "mean",
                "topk_sim_avg": "mean",
                "exact_score": "mean",
                "overlap_score": "mean",
                "semantic_score": "mean",
            })
            .rename(columns={
                "rr": "MRR",
                "hit_at_k": "Recall@K",
                "text_hit@k": "TextHit@K",
                "cluster_hit@k": "ClusterRecall@K",
                "rank": "RankAvg",
                "ndcg@k": "nDCG@K",
                "topk_sim_max": "TopKSimMaxAvg",
                "topk_sim_avg": "TopKSimAvg",
                "exact_score": "ExactScoreAvg",
                "overlap_score": "OverlapScoreAvg",
                "semantic_score": "SemanticScoreAvg",
            })
            .reset_index()
        )
        logger.info(f"Bloom summary shape: {bloom_summary.shape}")

    else:
        logger.info("BloomLevel column not found, skipping bloom analysis")
        bloom_summary = pd.DataFrame()  # fallback if missing

    #return df_res, chunks, summary, full_sim_matrix
    return df_res, chunks, summary, bloom_summary, full_sim_matrix, index, chunk_embs




# ================================================
# VISUALIZATION
# ================================================
def plot_heatmap(sim_matrix, out_dir, exp_id):
    logger.info(f"Plotting heatmap for {exp_id}")
    plt.figure(figsize=(10, 6))
    sns.heatmap(sim_matrix, cmap="viridis")
    plt.title(f"Similarity Heatmap: {exp_id}")
    path = f"{out_dir}/heatmap.png"
    plt.tight_layout()
    plt.savefig(path)
    plt.close()
    logger.info(f"Heatmap saved to {path}")
    return path


def plot_scatter_rank_similarity(df_res, out_dir):
    logger.info("Plotting scatter rank similarity")
    plt.figure(figsize=(6, 5))
    plt.scatter(df_res["rank"], df_res["topk_sim_max"], alpha=0.7)
    plt.xlabel("Rank")
    plt.ylabel("TopK Similarity Max")
    plt.title("Rank vs Similarity")
    plt.grid(True)

    path = f"{out_dir}/scatter_rank_similarity.png"
    plt.tight_layout()
    plt.savefig(path)
    plt.close()
    logger.info(f"Scatter plot saved to {path}")
    return path


# ================================================
# XLSX EXPORT
# ================================================
def save_excel(exp_id, out_dir, df_res, chunks, summary, bloom_summary, heatmap_path, scatter_path):
    logger.info(f"Saving Excel for {exp_id}")

    excel_path = f"{out_dir}/{exp_id}.xlsx"
    logger.info(f"Excel path: {excel_path}")

    df_summary = pd.DataFrame({
        "metric": summary.keys(),
        "value": summary.values()
    })
    logger.info(f"Summary DataFrame shape: {df_summary.shape}")

    df_chunks = pd.DataFrame({
        "chunk_id": range(1, len(chunks)+1),
        "chunk_text": chunks,
        "tokens": [len(c.split()) for c in chunks],
        "chars": [len(c) for c in chunks]
    })
    logger.info(f"Chunks DataFrame shape: {df_chunks.shape}")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as w:
        logger.info("Writing retrieval sheet...")
        df_res.to_excel(w, index=False, sheet_name="retrieval")
        ws = w.sheets["retrieval"]
        format_decimal(ws)

        logger.info("Writing summary sheet...")
        df_summary.to_excel(w, index=False, sheet_name="summary")
        ws2 = w.sheets["summary"]
        format_decimal(ws2)

        # ====================================================
        # Bloom Level Analysis Sheet
        # ====================================================
        if bloom_summary is not None and len(bloom_summary) > 0:
            logger.info("Writing bloom analysis sheet...")
            bloom_summary.to_excel(w, index=False, sheet_name="bloom_analysis")
            ws3 = w.sheets["bloom_analysis"]
            format_decimal(ws3)

        logger.info("Writing chunks sheet...")
        df_chunks.to_excel(w, index=False, sheet_name="chunks")
        ws4 = w.sheets["chunks"]
        format_decimal(ws4)

        wb = w.book

        logger.info("Adding heatmap image...")
        ws_h = wb.create_sheet("heatmap")
        img_h = openpyxl.drawing.image.Image(heatmap_path)
        img_h.anchor = "A1"
        ws_h.add_image(img_h)

        logger.info("Adding scatter plot image...")
        ws_s = wb.create_sheet("rank_similarity")
        img_s = openpyxl.drawing.image.Image(scatter_path)
        img_s.anchor = "A1"
        ws_s.add_image(img_s)

    logger.info(f"✔ Excel saved: {excel_path}")

# ======================================================
# SAVE FAISS + EMBEDDINGS + CHUNKS
# ======================================================
def save_faiss_and_metadata(out_dir, exp_id, index, chunk_embs, chunks):
    logger.info(f"Saving FAISS and metadata for {exp_id}")
    # Folder khusus FAISS
    faiss_dir = f"{out_dir}/faiss"
    os.makedirs(faiss_dir, exist_ok=True)
    logger.info(f"FAISS dir: {faiss_dir}")

    # Save FAISS index
    faiss_path = f"{faiss_dir}/index.faiss"
    faiss.write_index(index, faiss_path)
    logger.info(f"Saved FAISS index to {faiss_path}")

    # Save embeddings
    emb_path = f"{faiss_dir}/chunk_embs.npy"
    np.save(emb_path, chunk_embs)
    logger.info(f"Saved embeddings to {emb_path}, shape: {chunk_embs.shape}")

    # Save chunk texts (jsonl)
    chunks_path = f"{faiss_dir}/chunks.jsonl"
    with open(chunks_path, "w", encoding="utf8") as f:
        for i, ch in enumerate(chunks, start=1):
            rec = {"chunk_id": i, "text": ch}
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    logger.info(f"Saved chunks to {chunks_path}, {len(chunks)} chunks")

   

import ast

import ast

def flatten_context(x):
    try:
        d = ast.literal_eval(x)
    except:
        return x  # bukan dictionary / parsing gagal

    # recursive flatten
    def extract_values(obj):
        if isinstance(obj, dict):
            values = []
            for v in obj.values():
                values.append(extract_values(v))
            return " ".join(values)
        elif isinstance(obj, list):
            return " ".join(extract_values(i) for i in obj)
        else:
            return str(obj).strip()

    return extract_values(d)

# ================================================
# MAIN
# ================================================
def main():
    logger.info("Starting main function")
    cfg = load_yaml("configs/retrieval.yaml")
    logger.info(f"Loaded config from configs/retrieval.yaml")
    master = cfg["master"]
    experiments = cfg["experiments"]
    logger.info(f"Master config: {master}")
    logger.info(f"Number of experiments: {len(experiments)}")
    all_bloom = []

    df = pd.read_excel(master["dataset"])
    df = df[df["context"].str.strip() != ""]

    logger.info(f"Loaded dataset from {master['dataset']}, shape: {df.shape}")
    #df= df.head(10)  # for testing only, remove this line for full dataset
    #df["atomic_context_groundtruth"] = df["atomic_context_groundtruth"].astype(str).apply(flatten_context)
    #logger.info("Applied flatten_context to context column")


    ensure_dir(OUTPUT_DIR)
    logger.info(f"Ensured output directory: {OUTPUT_DIR}")

    global_summary = {}

    for exp in experiments:
        logger.info(f"Processing experiment: {exp['id']}")
        if not exp.get("active", False):
            logger.info(f"Experiment {exp['id']} is not active, skipping")
            continue
        logger.info(f"Experiment {exp['id']} is active, proceeding")

        exp_id = exp["id"]
        out_dir = auto_out(exp_id)
        logger.info(f"Output directory for {exp_id}: {out_dir}")

        df_res, chunks, summary, bloom_summary, sim_matrix, index, chunk_embs = run_retrieval(exp, master, df, out_dir)
        logger.info(f"Completed run_retrieval for {exp_id}, summary: {summary}")

        global_summary[exp_id] = summary
        bloom_summary["experiment_id"] = exp_id
        all_bloom.append(bloom_summary)
        logger.info(f"Added bloom_summary for {exp_id}")

        heatmap_path = plot_heatmap(sim_matrix, out_dir, exp_id)
        logger.info(f"Generated heatmap: {heatmap_path}")
        scatter_path = plot_scatter_rank_similarity(df_res, out_dir)
        logger.info(f"Generated scatter plot: {scatter_path}")

        # Save FAISS + embeddings + chunk metadata
        save_faiss_and_metadata(out_dir, exp_id, index, chunk_embs, chunks)
        logger.info(f"Saved FAISS and metadata for {exp_id}")

        save_excel(exp_id, out_dir, df_res, chunks, summary, bloom_summary, heatmap_path, scatter_path)
        logger.info(f"Saved Excel for {exp_id}")



        # SAVE GLOBAL COMPARISON
    summary_dir = OUTPUT_DIR
    ensure_dir(summary_dir)
    df_bloom_global = pd.concat(all_bloom, ignore_index=True)

    df_bloom_global.to_excel(
        f"{summary_dir}/summary_bloom_all_experiments_{stamp}.xlsx",
        index=False
    )
    df_global = pd.DataFrame.from_dict(global_summary, orient="index")
    df_global.reset_index(inplace=True)
    df_global.rename(columns={"index": "experiment_id"}, inplace=True)



    df_global.to_excel(
        f"{summary_dir}/summary_all_experiments_{stamp}.xlsx",
        index=False
    )




    logger.info("\n=== GLOBAL SUMMARY SAVED ===")

    logger.info("\n=== ALL DONE ===")


if __name__ == "__main__":
    main()
